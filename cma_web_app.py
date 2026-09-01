from __future__ import annotations

"""Local browser interface for the CMA dual-engine OCR package.

The app only binds preview links to 127.0.0.1. It never uploads PDFs or results.
OCR itself remains a separate child process, so closing a browser tab does not
stop a running job while this Streamlit service window stays open.
"""

import csv
import json
import os
import re
import shlex
import subprocess
import sys
import threading
import time
import unicodedata
from datetime import datetime
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, urlparse

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    from streamlit_autorefresh import st_autorefresh
except ImportError:
    # The packaged environment includes this small component. Keeping a safe
    # fallback lets the page open even if a manual installation omitted it.
    def st_autorefresh(**_kwargs: object) -> None:
        return


PACKAGE_DIR = Path(__file__).resolve().parent
OCR_SCRIPT = PACKAGE_DIR / "cma_dual_keyword_enhanced.py"
JOB_FILE_NAME = ".cma_web_job.json"
RAW_HEADER = [
    "文件名", "页码", "关键词", "实际匹配文本", "匹配方式", "匹配得分",
    "上下文", "识别引擎", "文本来源",
]


def _inside_root(root: Path, relative_name: str) -> Path | None:
    """Resolve a PDF only if it remains under the user-selected input root."""
    try:
        root = root.resolve()
        candidate = (root / relative_name).resolve()
        candidate.relative_to(root)
    except (OSError, ValueError):
        return None
    if candidate.suffix.casefold() != ".pdf" or not candidate.is_file():
        return None
    return candidate


class LocalPdfPreviewServer(ThreadingHTTPServer):
    allow_reuse_address = True
    daemon_threads = True

    def __init__(self, address: tuple[str, int]) -> None:
        super().__init__(address, LocalPdfRequestHandler)
        self.input_root = Path.cwd()


class LocalPdfRequestHandler(BaseHTTPRequestHandler):
    server: LocalPdfPreviewServer

    def log_message(self, _format: str, *_args: object) -> None:
        # Browser PDF range requests are normal; do not print them to the console.
        return

    def _requested_pdf(self) -> Path | None:
        parsed = urlparse(self.path)
        relative_name = parse_qs(parsed.query).get("file", [""])[0]
        return _inside_root(self.server.input_root, relative_name)

    def _send_text(self, status: HTTPStatus, text: str) -> None:
        data = text.encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self) -> None:  # noqa: N802 - required by BaseHTTPRequestHandler
        parsed = urlparse(self.path)
        pdf_path = self._requested_pdf()
        if pdf_path is None:
            self._send_text(HTTPStatus.NOT_FOUND, "<h3>找不到允许打开的 PDF 文件。</h3>")
            return

        if parsed.path == "/open":
            try:
                os.startfile(str(pdf_path))  # type: ignore[attr-defined]  # Windows only
            except OSError as exc:
                self._send_text(HTTPStatus.INTERNAL_SERVER_ERROR, f"<h3>打开失败：{exc}</h3>")
                return
            self._send_text(
                HTTPStatus.OK,
                "<h3>已交给 Windows 默认 PDF 阅读器打开。</h3><p>此标签页可关闭。</p>",
            )
            return

        if parsed.path != "/pdf":
            self._send_text(HTTPStatus.NOT_FOUND, "<h3>无效请求。</h3>")
            return

        try:
            size = pdf_path.stat().st_size
            start = 0
            end = size - 1
            range_header = self.headers.get("Range", "")
            match = re.match(r"bytes=(\d+)-(\d*)", range_header)
            if match:
                start = min(int(match.group(1)), end)
                if match.group(2):
                    end = min(int(match.group(2)), end)
                if end < start:
                    end = start
                self.send_response(HTTPStatus.PARTIAL_CONTENT)
                self.send_header("Content-Range", f"bytes {start}-{end}/{size}")
            else:
                self.send_response(HTTPStatus.OK)
            length = end - start + 1
            self.send_header("Content-Type", "application/pdf")
            self.send_header("Accept-Ranges", "bytes")
            self.send_header("Content-Length", str(length))
            # HTTP headers are latin-1; retain Chinese names through RFC 5987.
            fallback_name = pdf_path.name.encode("ascii", "replace").decode("ascii")
            utf8_name = quote(pdf_path.name, safe="")
            self.send_header(
                "Content-Disposition",
                f"inline; filename=\"{fallback_name}\"; filename*=UTF-8''{utf8_name}",
            )
            self.end_headers()
            with pdf_path.open("rb") as source:
                source.seek(start)
                remaining = length
                while remaining > 0:
                    chunk = source.read(min(1024 * 1024, remaining))
                    if not chunk:
                        break
                    self.wfile.write(chunk)
                    remaining -= len(chunk)
        except (OSError, ConnectionError):
            return


@st.cache_resource(show_spinner=False)
def get_preview_server() -> LocalPdfPreviewServer:
    for port in range(8765, 8786):
        try:
            server = LocalPdfPreviewServer(("127.0.0.1", port))
            threading.Thread(target=server.serve_forever, daemon=True).start()
            return server
        except OSError:
            continue
    raise RuntimeError("无法启动本地 PDF 预览服务（8765-8785 端口均不可用）。")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    try:
        with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as file:
            return [dict(row) for row in csv.DictReader(file) if row]
    except OSError:
        return []


def read_tail(path: Path, lines: int = 180) -> str:
    if not path.is_file():
        return "尚未生成日志。"
    try:
        raw = path.read_bytes()
        # cma_results.log is UTF-8. The fallback keeps manually generated
        # Windows console logs readable too, instead of filling the UI with �.
        try:
            content = raw.decode("utf-8").splitlines()
        except UnicodeDecodeError:
            content = raw.decode("gb18030", errors="replace").splitlines()
        return "\n".join(content[-lines:]) or "日志目前为空。"
    except OSError as exc:
        return f"无法读取日志：{exc}"


def read_job(output_root: Path) -> dict[str, Any]:
    path = output_root / JOB_FILE_NAME
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}


def write_job(output_root: Path, job: dict[str, Any]) -> None:
    output_root.mkdir(parents=True, exist_ok=True)
    (output_root / JOB_FILE_NAME).write_text(
        json.dumps(job, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def process_is_running(pid: int | None) -> bool:
    if not pid or pid <= 0:
        return False
    try:
        result = subprocess.run(
            ["tasklist", "/FI", f"PID eq {pid}", "/NH"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
        )
    except OSError:
        return False
    return str(pid) in result.stdout


def read_resume_state(output_root: Path) -> dict[str, Any]:
    try:
        return json.loads(
            (output_root / ".cma_ocr_resume" / "state.json").read_text(encoding="utf-8")
        )
    except (OSError, ValueError, TypeError):
        return {}


def state_matches_job(state: dict[str, Any], job: dict[str, Any]) -> bool:
    """Reject raw files left by an older task during the new process startup gap."""
    if not state or not job:
        return False
    try:
        return (
            Path(str(state.get("input_dir", ""))).resolve()
            == Path(str(job.get("input_root", ""))).resolve()
            and state.get("keywords") == job.get("keyword_list")
            and int(state.get("dpi", 0)) == int(job.get("dpi", 0))
        )
    except (OSError, TypeError, ValueError):
        return False


def latest_run_dir(
    output_root: Path, job: dict[str, Any], is_running: bool
) -> Path | None:
    pointer = output_root / "CMA_OCR_最近一次结果.txt"
    try:
        # Immediately after a new click the pointer may still name the previous
        # run. Do not flash old merged results while the new child is starting.
        started_epoch = float(job.get("started_epoch", 0) or 0)
        if is_running and started_epoch and pointer.stat().st_mtime + 2 < started_epoch:
            return None
        candidate = Path(pointer.read_text(encoding="utf-8").strip()).resolve()
        candidate.relative_to((output_root / "runs").resolve())
        return candidate if candidate.is_dir() else None
    except (OSError, ValueError, TypeError):
        return None


def active_log_path(
    output_root: Path, job: dict[str, Any], is_running: bool
) -> Path:
    """Use the run log because it also contains Paddle child page-level events."""
    run_dir = latest_run_dir(output_root, job, is_running)
    if run_dir is not None:
        run_log = run_dir / "cma_results.log"
        if run_log.is_file():
            return run_log
    return output_root / "cma_results.log"


def paddle_worker_log_path(run_dir: Path | None) -> Path | None:
    if run_dir is None:
        return None
    path = run_dir / "cma_results_paddle_worker.log"
    return path if path.is_file() else None


def read_paddle_worker_status(output_root: Path, is_running: bool) -> dict[str, Any]:
    if not is_running:
        return {}
    path = output_root / ".cma_ocr_resume" / "paddle_worker_status.json"
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return {}


def _score_number(row: dict[str, str]) -> float:
    try:
        return float(row.get("匹配得分", "0") or 0)
    except ValueError:
        return 0.0


def provisional_rows(rows: list[dict[str, str]], engine: str) -> list[dict[str, str]]:
    """Collapse repeated occurrences without hiding distinct page/keyword hits."""
    method_rank = {"exact": 3, "ocr_confusion": 2, "fuzzy": 1}
    best: dict[tuple[str, str, str, str], dict[str, str]] = {}
    for source in rows:
        row = dict(source)
        row_engine = row.get("识别引擎", engine) or engine
        row["来源引擎"] = row_engine
        row["引擎结论"] = "单引擎阶段结果"
        row["结果状态"] = "阶段性（双引擎尚未合并）"
        row["置信度"] = "阶段性"
        key = (
            row.get("文件名", ""), row.get("页码", ""),
            row.get("关键词", ""), row_engine,
        )
        old = best.get(key)
        rank = (method_rank.get(row.get("匹配方式", ""), 0), _score_number(row))
        old_rank = (
            method_rank.get(old.get("匹配方式", ""), 0), _score_number(old)
        ) if old else (-1, -1.0)
        if old is None or rank > old_rank:
            best[key] = row
    return sorted(
        best.values(),
        key=lambda row: (row.get("文件名", ""), int(float(row.get("页码", "0") or 0)), row.get("关键词", "")),
    )


def merged_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    decorated: list[dict[str, str]] = []
    for source in rows:
        row = dict(source)
        engines = {
            item.strip().casefold()
            for item in row.get("来源引擎", "").split("+")
            if item.strip()
        }
        row["引擎结论"] = "双引擎共同命中" if len(engines) >= 2 else "单引擎命中"
        row["结果状态"] = "最终 merge 去重结果"
        decorated.append(row)
    return decorated


def current_result_view(
    output_root: Path,
    job: dict[str, Any],
    is_running: bool,
    state: dict[str, Any],
) -> tuple[list[dict[str, str]], str, bool, Path | None]:
    """Show engine-1 provisionally, then atomically switch to final merge output."""
    run_dir = latest_run_dir(output_root, job, is_running)
    if run_dir is not None:
        final_path = run_dir / "cma_results_merged.csv"
        if final_path.is_file():
            return merged_rows(read_csv_rows(final_path)), "最终 merge 结果", True, run_dir

    if not state_matches_job(state, job):
        return [], "等待本次任务写入首个结果", False, run_dir

    state_dir = output_root / ".cma_ocr_resume"
    t_done = len(state.get("tesseract_done", []))
    p_done = len(state.get("paddle_done", []))
    t_path = state_dir / "tesseract_raw.csv"
    p_path = state_dir / "paddle_raw.csv"
    # The normal order is Tesseract then Paddle. Even while Paddle gradually
    # produces raw hits, keep the completed first-engine view stable until the
    # final merge file appears, as requested.
    if t_done or t_path.is_file():
        return provisional_rows(read_csv_rows(t_path), "tesseract"), (
            "Tesseract 阶段结果（非最终，等待双引擎 merge）"
        ), False, run_dir
    if p_done or p_path.is_file():
        return provisional_rows(read_csv_rows(p_path), "paddle"), (
            "PaddleOCR 阶段结果（非最终，等待双引擎 merge）"
        ), False, run_dir
    return [], "等待首个引擎产生结果", False, run_dir


def filter_rows(
    rows: list[dict[str, str]],
    filename_query: str,
    selected_keywords: list[str],
    selected_engines: list[str],
    selected_confidences: list[str],
) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    needle = filename_query.strip().casefold()
    for row in rows:
        if needle and needle not in row.get("文件名", "").casefold():
            continue
        if selected_keywords and row.get("关键词", "") not in selected_keywords:
            continue
        if selected_engines and row.get("识别引擎", row.get("来源引擎", "")) not in selected_engines:
            continue
        if selected_confidences and row.get("置信度", "") not in selected_confidences:
            continue
        result.append(row)
    return result


def write_sheet(book: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    sheet = book.create_sheet(title)
    headers = list(dict.fromkeys(key for row in rows for key in row))
    if not headers:
        sheet.append(["无数据"])
        return
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        width = min(55, max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def make_xlsx(
    filtered_rows: list[dict[str, str]],
    trusted_rows: list[dict[str, str]],
    candidate_rows: list[dict[str, str]],
    error_rows: list[dict[str, str]],
    summary: dict[str, str],
) -> bytes:
    book = Workbook()
    book.remove(book.active)
    write_sheet(book, "当前筛选结果", filtered_rows)
    write_sheet(book, "可信命中", trusted_rows)
    write_sheet(book, "待复核候选", candidate_rows)
    write_sheet(book, "错误明细", error_rows)
    write_sheet(book, "运行摘要", [{"项目": key, "数值": value} for key, value in summary.items()])
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def latest_status_line(log_text: str) -> str:
    useful = [line for line in log_text.splitlines() if "|" in line]
    return useful[-1] if useful else "等待任务启动。"


def parse_progress(
    log_text: str,
    resume_state: dict[str, Any] | None = None,
    paddle_status: dict[str, Any] | None = None,
) -> dict[str, str | int | float | None]:
    """Extract the same file/page/heartbeat signals used by the desktop app.

    The OCR log is intentionally detailed for troubleshooting.  This parser
    presents only its latest useful state in the web page, without changing
    the OCR process or its original log file.
    """
    result: dict[str, str | int | float | None] = {
        "engine": None,
        "phase": None,
        "filename": None,
        "file_index": None,
        "file_total": None,
        "page": None,
        "page_total": None,
        "heartbeat": None,
        "elapsed": None,
        "heartbeat_scope": None,
        "worker_stage": None,
        "event": None,
    }
    phase_pattern = re.compile(r"==========\s+(tesseract|paddle)\s+独立识别阶段开始\s+==========", re.I)
    paddle_launch_pattern = re.compile(r"启动 Paddle 子进程处理本批", re.I)
    paddle_created_pattern = re.compile(r"Paddle 子进程已创建\s+\|\s+PID=", re.I)
    paddle_batch_pattern = re.compile(r"==========\s+Paddle 子进程启动", re.I)
    file_pattern = re.compile(r"\[(\d+)/(\d+)\]\s+(?!跳过（已完成）)(.+)$")
    page_pattern = re.compile(
        r"\[开始页面\]\s+引擎=([^|]+)\s+\|\s+文件=(.+?)\s+\|\s+"
        r"文件进度=(\d+)/(\d+)\s+\|\s+页码=(\d+)/(\d+)"
    )
    heartbeat_pattern = re.compile(
        r"\[心跳#(\d+)\s+仍在识别\]\s+引擎=([^|]+)\s+\|\s+"
        r"文件=(.+?)\s+\|\s+页码=(\d+)\s+\|\s+已耗时=([\d.]+)s"
    )
    worker_heartbeat_pattern = re.compile(
        r"\[Paddle 子进程心跳#(\d+)\]\s+PID=\d+\s+\|\s+阶段=(.+)$"
    )

    def switch_engine(engine: str, event: str, phase: str | None = None) -> None:
        # A model phase is an independent event. Clear the previous model's
        # filename/page/heartbeat so Paddle never inherits Tesseract's 100%.
        result.update({
            "engine": engine,
            "phase": phase or engine,
            "filename": None,
            "file_index": 0,
            "file_total": None,
            "page": None,
            "page_total": None,
            "heartbeat": None,
            "elapsed": None,
            "heartbeat_scope": None,
            "worker_stage": None,
            "event": event,
        })

    for line in log_text.splitlines():
        phase_match = phase_pattern.search(line)
        if phase_match:
            engine_name = phase_match.group(1).lower()
            switch_engine(engine_name, line)
            continue

        if paddle_launch_pattern.search(line):
            switch_engine("paddle", line, "Paddle 子进程准备")
            continue

        if paddle_created_pattern.search(line):
            if result["engine"] != "paddle":
                switch_engine("paddle", line, "Paddle 子进程启动")
            else:
                result["phase"] = "Paddle 子进程启动"
                result["event"] = line
            continue

        if paddle_batch_pattern.search(line):
            if result["engine"] != "paddle":
                switch_engine("paddle", line, "PaddleOCR 识别")
            else:
                result["phase"] = "PaddleOCR 识别"
                result["event"] = line
            continue

        page_match = page_pattern.search(line)
        if page_match:
            result.update({
                "engine": page_match.group(1).strip(),
                "filename": page_match.group(2).strip(),
                "file_index": int(page_match.group(3)),
                "file_total": int(page_match.group(4)),
                "page": int(page_match.group(5)),
                "page_total": int(page_match.group(6)),
            })
            continue

        heartbeat_match = heartbeat_pattern.search(line)
        if heartbeat_match:
            result.update({
                "engine": heartbeat_match.group(2).strip(),
                "filename": heartbeat_match.group(3).strip(),
                "page": int(heartbeat_match.group(4)),
                "heartbeat": int(heartbeat_match.group(1)),
                "elapsed": float(heartbeat_match.group(5)),
                "heartbeat_scope": "页面",
                "event": line,
            })
            continue

        worker_heartbeat_match = worker_heartbeat_pattern.search(line)
        if worker_heartbeat_match:
            result.update({
                "engine": "paddle",
                "phase": "Paddle 子进程",
                "heartbeat": int(worker_heartbeat_match.group(1)),
                "heartbeat_scope": "子进程",
                "worker_stage": worker_heartbeat_match.group(2).strip(),
                "event": line,
            })
            continue

        file_match = file_pattern.search(line)
        if file_match and " | INFO | " in line:
            result.update({
                "file_index": int(file_match.group(1)),
                "file_total": int(file_match.group(2)),
                "filename": file_match.group(3).strip(),
                "event": line,
            })
            continue

        # Keep a recent meaningful event, but deliberately exclude the
        # one-line-per-page DEBUG trace that made the previous view too busy.
        if " | DEBUG | " not in line and (
            "[心跳" in line
            or "[Paddle 子进程" in line
            or "[页面无结果]" in line
            or "[内存回退]" in line
            or "独立识别阶段开始" in line
            or "识别完成" in line
        ):
            result["event"] = line

    worker = paddle_status or {}
    if worker:
        if result.get("engine") != "paddle":
            switch_engine("paddle", "Paddle 子进程已报告实时状态", "PaddleOCR 识别")
        result["phase"] = "PaddleOCR 识别"
        result["worker_stage"] = str(worker.get("stage", "") or "Paddle 子进程运行中")
        if worker.get("filename"):
            result["filename"] = str(worker["filename"])
            for key in ("page", "page_total"):
                try:
                    result[key] = int(worker.get(key, 0)) or None
                except (TypeError, ValueError):
                    result[key] = None
        else:
            result["filename"] = None
            result["page"] = None
            result["page_total"] = None

    state = resume_state or {}
    manifest = state.get("files", [])
    total_files = len(manifest) if isinstance(manifest, list) else 0
    engine_name = str(result.get("engine") or "").casefold()
    if total_files and engine_name == "paddle":
        completed = len(state.get("paddle_done", []))
        result["file_total"] = total_files
        # Child logs use batch-local positions. Resume state is global and is
        # updated after every completed file, so completed+1 is the active file.
        if result.get("filename") and result.get("page"):
            result["file_index"] = min(total_files, completed + 1)
        else:
            result["file_index"] = min(total_files, completed)
    elif total_files and engine_name == "tesseract":
        result["file_total"] = total_files
    return result


def short_event(text: str | None, limit: int = 180) -> str:
    if not text:
        return "等待 OCR 写入第一条进度信息。"
    return text if len(text) <= limit else f"{text[:limit - 1]}…"


def parsed_keyword_list(raw: str) -> list[str]:
    if re.search(r"[,，、;；]", raw):
        parts = re.split(r"[,，、;；]+", raw)
    else:
        parts = shlex.split(raw, posix=False)
    keywords: list[str] = []
    seen: set[str] = set()
    for part in parts:
        cleaned = part.strip().strip('"\'')
        identity = re.sub(r"\s+", "", unicodedata.normalize("NFKC", cleaned)).casefold()
        if cleaned and identity and identity not in seen:
            seen.add(identity)
            keywords.append(cleaned)
    return keywords or ["CMA"]


def start_ocr(input_root: Path, output_root: Path, keywords: str, dpi: int, binarize: bool) -> None:
    command = [
        sys.executable, str(OCR_SCRIPT), str(input_root), "-k", keywords,
        "-o", str(output_root), "--dpi", str(dpi),
    ]
    if binarize:
        command.append("--binarize")
    child_env = os.environ.copy()
    child_env["PYTHONUTF8"] = "1"
    child_env["PYTHONIOENCODING"] = "utf-8"
    # Inherit the launcher CMD window.  The full desktop-style file progress,
    # page progress, heartbeat and error lines remain visible there; the web
    # page is deliberately only the compact companion view.
    process = subprocess.Popen(
        command,
        cwd=str(PACKAGE_DIR),
        creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
        env=child_env,
    )
    write_job(
        output_root,
        {
            "pid": process.pid,
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "input_root": str(input_root),
            "output_root": str(output_root),
            "keywords": keywords,
            "keyword_list": parsed_keyword_list(keywords),
            "dpi": dpi,
            "binarize": binarize,
            "started_epoch": time.time(),
        },
    )


def main() -> None:
    st.title("CMA 双模型 OCR · 本地网页版")
    st.caption("PDF 和识别结果仅保留在本机；关闭浏览器标签页不会停止后台 OCR。")

    with st.sidebar:
        st.header("任务设置")
        keywords = st.text_input("关键词", value="CMA,资质,检测")
        input_text = st.text_input("PDF 根文件夹", placeholder=r"D:\PDF")
        output_text = st.text_input("结果保存文件夹", placeholder=r"D:\OCR_Results")
        dpi = st.number_input("渲染 DPI", min_value=200, max_value=600, value=400, step=25)
        binarize = st.checkbox("Tesseract 二值化", value=False)
        auto_refresh = st.checkbox("任务运行时自动更新（每 10 秒）", value=True)
        start_clicked = st.button("开始后台 OCR", type="primary", use_container_width=True)
        st.button("立即刷新当前页面", use_container_width=True)

    if start_clicked:
        input_root = Path(input_text.strip().strip('"')).expanduser()
        output_root = Path(output_text.strip().strip('"')).expanduser()
        if not keywords.strip():
            st.error("请至少输入一个关键词。")
        elif not input_root.is_dir():
            st.error("PDF 根文件夹无效。")
        elif not output_text.strip():
            st.error("请填写结果保存文件夹。")
        else:
            output_root.mkdir(parents=True, exist_ok=True)
            previous = read_job(output_root)
            if process_is_running(previous.get("pid")):
                st.warning("该结果目录已有 OCR 在运行；为防止重复占用内存，未启动第二个任务。")
            else:
                start_ocr(input_root.resolve(), output_root.resolve(), keywords.strip(), int(dpi), binarize)
                st.success("后台 OCR 已启动。保持本启动窗口开启即可；可关闭或刷新浏览器页面。")

    active_output = output_text.strip().strip('"')
    if not active_output:
        st.info("填写“结果保存文件夹”后，可在此实时查看本次或历史任务的结果。")
        return
    output_root = Path(active_output).expanduser()
    job = read_job(output_root)
    is_running = process_is_running(job.get("pid"))
    resume_state = read_resume_state(output_root)
    log_path = active_log_path(output_root, job, is_running)
    log_text = read_tail(log_path, lines=500)
    display_rows, result_mode, has_final_merge, run_dir = current_result_view(
        output_root, job, is_running, resume_state
    )
    # While a new child is starting, root-level CSVs may still belong to the
    # previous task. An impossible placeholder keeps those stale files hidden.
    result_root = run_dir if run_dir is not None else (
        output_root if not is_running else output_root / ".waiting_for_current_run"
    )
    trusted_rows = display_rows if has_final_merge else []
    candidate_rows = read_csv_rows(result_root / "cma_results_candidates.csv")
    error_rows = read_csv_rows(result_root / "cma_results_errors.csv")
    paddle_status = read_paddle_worker_status(output_root, is_running)
    worker_log_path = paddle_worker_log_path(run_dir)
    worker_log_text = read_tail(worker_log_path, lines=300) if worker_log_path else ""

    # Do not refresh a finished/history view: filters and selected rows stay
    # still unless the user explicitly requests a refresh.
    if auto_refresh and is_running:
        st_autorefresh(interval=10000, key="cma_ocr_auto_refresh")

    status = "正在运行" if is_running else ("已结束或未启动" if job else "未发现任务记录")
    metric1, metric2, metric3, metric4 = st.columns(4)
    metric1.metric("任务状态", status)
    metric2.metric("当前展示命中", len(display_rows))
    metric3.metric("结果阶段", "最终 merge" if has_final_merge else "阶段性")
    metric4.metric("错误记录", len(error_rows))
    st.caption(f"当前数据源：{result_mode}")
    progress = parse_progress(log_text, resume_state, paddle_status)
    st.subheader("识别进展")
    progress_left, progress_mid, progress_right, progress_last = st.columns(4)
    engine_value = str(progress["engine"] or "").casefold()
    if engine_value == "tesseract":
        phase_name = "Tesseract（1 / 2）"
    elif engine_value == "paddle":
        phase_name = "PaddleOCR（2 / 2）"
    else:
        phase_name = str(progress["phase"] or "等待启动")
    progress_left.metric("当前阶段", phase_name)
    file_index = progress["file_index"]
    file_total = progress["file_total"]
    page = progress["page"]
    page_total = progress["page_total"]
    shown_file_index = file_index if isinstance(file_index, int) else "-"
    shown_file_total = file_total if isinstance(file_total, int) else "-"
    shown_page = page if isinstance(page, int) else "-"
    shown_page_total = page_total if isinstance(page_total, int) else "-"
    progress_mid.metric("文件进度", f"{shown_file_index} / {shown_file_total}")
    progress_right.metric("当前页", f"{shown_page} / {shown_page_total}")
    heartbeat = progress["heartbeat"]
    elapsed = progress["elapsed"]
    heartbeat_scope = str(progress["heartbeat_scope"] or "活动")
    if heartbeat is None:
        heartbeat_text = "等待心跳"
    elif isinstance(elapsed, (int, float)):
        heartbeat_text = f"{heartbeat_scope} #{heartbeat} · {elapsed:.0f} 秒"
    else:
        heartbeat_text = f"{heartbeat_scope} #{heartbeat}"
    progress_last.metric("活动心跳", heartbeat_text)
    worker_stage = str(progress.get("worker_stage") or "")
    filename = str(progress["filename"] or worker_stage or "等待任务开始")
    if isinstance(file_index, int) and isinstance(file_total, int) and file_total > 0:
        file_fraction = min(1.0, max(0.0, file_index / file_total))
        st.progress(file_fraction, text=f"{phase_name}｜文件进度 {file_index}/{file_total}")
    if isinstance(page, int) and isinstance(page_total, int) and page_total > 0:
        page_fraction = min(1.0, max(0.0, page / page_total))
        st.progress(page_fraction, text=f"当前文件：{filename} ｜第 {page}/{page_total} 页")
    else:
        st.info(f"当前文件：{filename}")
    st.caption(f"最近事件：{short_event(progress['event'] if isinstance(progress['event'], str) else None)}")
    if job:
        st.caption(
            f"开始时间：{job.get('started_at', '-')}　关键词：{job.get('keywords', '-')}　"
            f"PDF 根目录：{job.get('input_root', '-')}"
        )

    current_label = "当前命中（最终 merge）" if has_final_merge else "当前命中（阶段性）"
    source_choice = st.radio(
        "查看数据", [current_label, "最终可信命中", "待复核候选", "错误明细"], horizontal=True
    )
    source_rows = {
        current_label: display_rows,
        "最终可信命中": trusted_rows,
        "待复核候选": candidate_rows,
        "错误明细": error_rows,
    }[source_choice]

    all_keywords = sorted({row.get("关键词", "") for row in source_rows if row.get("关键词", "")})
    all_engines = sorted({
        row.get("识别引擎", row.get("来源引擎", ""))
        for row in source_rows if row.get("识别引擎", row.get("来源引擎", ""))
    })
    all_confidences = sorted({row.get("置信度", "") for row in source_rows if row.get("置信度", "")})
    filter_left, filter_mid, filter_right, filter_last = st.columns(4)
    with filter_left:
        filename_query = st.text_input("文件名包含", key="filename_filter")
    with filter_mid:
        selected_keywords = st.multiselect("关键词筛选", all_keywords)
    with filter_right:
        selected_engines = st.multiselect("模型来源筛选", all_engines)
    with filter_last:
        selected_confidences = st.multiselect("置信度筛选", all_confidences)

    filtered_rows = filter_rows(
        source_rows, filename_query, selected_keywords, selected_engines, selected_confidences
    )
    st.subheader(f"{source_choice}（{len(filtered_rows)} 条）")
    st.dataframe(filtered_rows, use_container_width=True, hide_index=True, height=380)

    pdf_rows = [row for row in filtered_rows if row.get("文件名") and row.get("页码")]
    if pdf_rows:
        selected_index = st.selectbox(
            "选择一条命中进行回查",
            options=list(range(len(pdf_rows))),
            format_func=lambda index: f"{pdf_rows[index].get('文件名')} ｜第{pdf_rows[index].get('页码')}页｜{pdf_rows[index].get('关键词', '')}",
        )
        selected = pdf_rows[selected_index]
        root_text = job.get("input_root", "") if job else input_text.strip().strip('"')
        source_root = Path(root_text).expanduser() if root_text else None
        if source_root and source_root.is_dir():
            preview_server = get_preview_server()
            preview_server.input_root = source_root.resolve()
            relative_name = selected["文件名"]
            try:
                page_number = max(1, int(float(selected["页码"])))
            except ValueError:
                page_number = 1
            encoded_name = quote(relative_name, safe="")
            preview_url = f"http://127.0.0.1:{preview_server.server_port}/pdf?file={encoded_name}#page={page_number}"
            open_url = f"http://127.0.0.1:{preview_server.server_port}/open?file={encoded_name}"
            action_left, action_mid, action_right = st.columns([1, 1, 2])
            with action_left:
                st.link_button("新标签预览该页", preview_url, use_container_width=True)
            with action_mid:
                st.link_button("默认阅读器打开", open_url, use_container_width=True)
            with action_right:
                st.code(str((source_root / relative_name).resolve()), language=None)
        else:
            st.warning("未找到本次任务的 PDF 根目录，暂不能打开对应文件。")

    summary = {
        "导出时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "当前视图": source_choice,
        "当前筛选条数": str(len(filtered_rows)),
        "当前展示命中": str(len(display_rows)),
        "结果阶段": result_mode,
        "最终可信命中": str(len(trusted_rows)),
        "待复核候选": str(len(candidate_rows)),
        "错误记录": str(len(error_rows)),
        "结果保存文件夹": str(output_root),
    }
    st.download_button(
        "导出当前结果为 XLSX",
        data=make_xlsx(filtered_rows, trusted_rows, candidate_rows, error_rows, summary),
        file_name=f"CMA_OCR_筛选结果_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    with st.expander("完整排错日志（按需展开）", expanded=False):
        st.caption("主进程 / Tesseract / Paddle 心跳日志")
        st.code(log_text, language="text")
        if worker_log_text:
            st.caption("PaddleOCR 子进程详细页级日志")
            st.code(worker_log_text, language="text")


if __name__ == "__main__":
    main()

