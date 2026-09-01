from __future__ import annotations

import argparse
import atexit
import csv
import gc
import hashlib
import json
import logging
import os
import queue
import re
import shlex
import shutil
import subprocess
import sys
import threading
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeoutError
from dataclasses import asdict, dataclass, field
from difflib import SequenceMatcher
from importlib import metadata
from pathlib import Path
from typing import Any, Iterable


# ======================== Isolated environment bootstrap ====================
# This section intentionally uses only the Python standard library. It runs
# before numpy/OpenCV/Paddle/PyMuPDF are imported, so a broken base environment
# can still create and launch the isolated OCR environment.
BOOTSTRAP_ENV_NAME = "cma_ocr27_web"
BOOTSTRAP_INDEX_URL = (
    os.environ.get("CMA_PIP_INDEX_URL")
    or os.environ.get("PIP_INDEX_URL")
    or "https://pypi.org/simple"
)
BOOTSTRAP_REQUIREMENTS = Path(__file__).with_name("requirements.txt")


def _same_path(left: Path, right: Path) -> bool:
    try:
        return str(left.resolve()).casefold() == str(right.resolve()).casefold()
    except OSError:
        return False


def _find_conda_executable() -> Path | None:
    candidates: list[Path] = []
    for variable_name in ("CMA_CONDA_EXE", "CONDA_EXE"):
        configured = os.environ.get(variable_name, "").strip().strip('"')
        if configured:
            candidates.append(Path(configured).expanduser())

    executable_dir = Path(sys.executable).resolve().parent
    home_dir = Path.home()
    program_data = Path(os.environ.get("ProgramData", r"C:\ProgramData"))
    candidates.extend([
        executable_dir / "Scripts" / "conda.exe",
        executable_dir.parent.parent / "Scripts" / "conda.exe",
        home_dir / "anaconda3" / "Scripts" / "conda.exe",
        home_dir / "miniconda3" / "Scripts" / "conda.exe",
        program_data / "anaconda3" / "Scripts" / "conda.exe",
        program_data / "miniconda3" / "Scripts" / "conda.exe",
    ])
    discovered = shutil.which("conda")
    if discovered:
        candidates.append(Path(discovered))

    checked: set[str] = set()
    for candidate in candidates:
        normalized = os.path.normcase(os.path.abspath(str(candidate)))
        if normalized in checked:
            continue
        checked.add(normalized)
        if candidate.is_file():
            return candidate
    return None


def _read_pinned_requirements(path: Path) -> dict[str, str]:
    expected: dict[str, str] = {}
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "==" not in line:
            raise ValueError(f"依赖必须固定版本：{line}")
        package, version = line.split("==", 1)
        expected[package.strip()] = version.strip()
    return expected


def _environment_healthcheck(
    env_python: Path,
    expected: dict[str, str],
) -> tuple[bool, str]:
    check_code = f"""
import importlib.metadata as metadata
import sys

expected = {expected!r}
problems = []
for package, wanted in expected.items():
    try:
        actual = metadata.version(package)
    except metadata.PackageNotFoundError:
        problems.append(f"{{package}}: missing")
        continue
    if actual != wanted:
        problems.append(f"{{package}}: {{actual}} != {{wanted}}")

if problems:
    print("; ".join(problems))
    sys.exit(2)

try:
    import numpy
    import cv2
    import paddle
    import fitz
    import pytesseract
    import openpyxl
    import setuptools
    from paddleocr import PaddleOCR
except Exception as exc:
    print(f"critical import failed: {{type(exc).__name__}}: {{exc}}")
    sys.exit(3)

print("dependency and ABI check passed")
"""
    completed = subprocess.run(
        [str(env_python), "-c", check_code],
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        check=False,
    )
    details = "\n".join(
        part.strip() for part in (completed.stdout, completed.stderr) if part.strip()
    )
    return completed.returncode == 0, details


def _bootstrap_isolated_environment() -> None:
    conda_executable = _find_conda_executable()
    if conda_executable is None:
        raise RuntimeError("未找到 conda.exe，无法创建独立 OCR 环境")
    if not BOOTSTRAP_REQUIREMENTS.is_file():
        raise FileNotFoundError(f"缺少依赖文件：{BOOTSTRAP_REQUIREMENTS}")

    conda_root = conda_executable.parent.parent
    env_dir = conda_root / "envs" / BOOTSTRAP_ENV_NAME
    env_python = env_dir / "python.exe"

    # Once re-launched inside the isolated interpreter, continue to OCR imports.
    if env_python.is_file() and _same_path(Path(sys.executable), env_python):
        return

    expected = _read_pinned_requirements(BOOTSTRAP_REQUIREMENTS)
    created_now = False
    if not env_python.is_file():
        print(f"[环境准备] 首次运行，正在创建独立环境：{BOOTSTRAP_ENV_NAME}")
        print("[环境准备] 此过程可能需要数分钟，请不要关闭窗口。")
        create_result = subprocess.run(
            [
                str(conda_executable),
                "create",
                "-p",
                str(env_dir),
                "python=3.10",
                "pip",
                "-y",
            ],
            check=False,
        )
        if create_result.returncode != 0 or not env_python.is_file():
            raise RuntimeError("独立环境创建失败，请保留当前窗口的完整输出")
        created_now = True

    healthy, details = _environment_healthcheck(env_python, expected)
    if healthy:
        print(f"[环境检查] {BOOTSTRAP_ENV_NAME} 依赖和 ABI 正常，无需安装。")
    else:
        print(f"[环境检查] 需要安装或修复依赖：{details or '首次安装'}")
        pip_command = [
            str(env_python),
            "-m",
            "pip",
            "install",
            "-r",
            str(BOOTSTRAP_REQUIREMENTS),
            "-i",
            BOOTSTRAP_INDEX_URL,
        ]
        # If exact versions exist but imports are broken, reinstall only inside
        # the isolated environment. The base environment remains untouched.
        if not created_now and "critical import failed" in details:
            pip_command.insert(4, "--force-reinstall")
        install_result = subprocess.run(pip_command, check=False)
        if install_result.returncode != 0:
            raise RuntimeError("OCR 依赖安装失败，请保留当前窗口的完整输出")
        healthy, details = _environment_healthcheck(env_python, expected)
        if not healthy:
            raise RuntimeError(f"依赖安装后检查仍未通过：{details}")
        print("[环境检查] 安装完成，依赖和 ABI 检查通过。")

    print(f"[环境启动] 正在使用 {env_python} 运行 OCR。")
    completed = subprocess.run(
        [str(env_python), str(Path(__file__).resolve()), *sys.argv[1:]],
        cwd=str(Path.cwd()),
        check=False,
    )
    raise SystemExit(completed.returncode)


try:
    _bootstrap_isolated_environment()
except SystemExit:
    raise
except Exception as bootstrap_error:
    print(f"[环境错误] {type(bootstrap_error).__name__}: {bootstrap_error}")
    print("请不要在 base 环境继续安装包；保存完整输出用于排查。")
    raise SystemExit(1)
# ===========================================================================


# Paddle 以独立子进程运行。状态文件让父进程能在“尚未写入常规日志”的导入或
# 初始化阶段也持续给出心跳，避免界面误把正常加载误判成卡死。
PADDLE_WORKER_MODE = len(sys.argv) >= 3 and sys.argv[1] == "--_paddle_worker"
PADDLE_WORKER_STATUS_FILE = os.environ.get("CMA_PADDLE_WORKER_STATUS_FILE", "")


def _report_paddle_worker_stage(stage: str, **details: Any) -> None:
    if not PADDLE_WORKER_MODE:
        return
    message = {"pid": os.getpid(), "stage": stage, "updated_at": time.time()}
    message.update(details)
    if PADDLE_WORKER_STATUS_FILE:
        target = Path(PADDLE_WORKER_STATUS_FILE)
        temporary = target.with_suffix(target.suffix + ".tmp")
        try:
            target.parent.mkdir(parents=True, exist_ok=True)
            temporary.write_text(json.dumps(message, ensure_ascii=False), encoding="utf-8")
            temporary.replace(target)
        except OSError:
            pass
    print(f"[Paddle 子进程] {stage}", flush=True)


_report_paddle_worker_stage("隔离解释器已启动，正在导入 OCR 依赖")

# These flags must be set before PaddleOCR is imported.
os.environ.setdefault("FLAGS_use_mkldnn", "0")
os.environ.setdefault("FLAGS_use_onednn", "0")

import numpy as np
try:
    import pymupdf as fitz
except ImportError:
    # PyMuPDF 1.20.x (required by PaddleOCR 2.7.0.3) exposes the old name.
    import fitz
import pytesseract
from PIL import Image, ImageFilter, ImageOps

_report_paddle_worker_stage("基础依赖导入完成，等待初始化 PaddleOCR")


# ============================== Configuration ==============================
# Tesseract 支持显式环境变量、Windows 常见安装位置以及 PATH 自动查找。
TESSERACT_EXE_CANDIDATES = [
    *(
        [Path(os.environ["TESSERACT_CMD"]).expanduser()]
        if os.environ.get("TESSERACT_CMD")
        else []
    ),
    Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
    Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
]
TESS_LANG = "chi_sim+eng"
TESS_CONFIG = "--oem 3 --psm 3 -c preserve_interword_spaces=1"
PADDLE_LANG = "ch"

# Both engines always OCR every page. The PDF text layer is only used if an OCR
# engine returns too little usable text. This keeps the two OCR runs independent.
RENDER_DPI = 400  # 扫描件在 300 下易被压糊，400 更清晰
BINARIZE = False  # 扫描件/浅字可设为 True；清晰数字版 PDF 建议保持 False
MIN_OCR_TEXT_LENGTH = 8
USE_PDF_TEXT_AS_FALLBACK = True
MAX_RETRIES = 5
RETRY_DELAY_SECONDS = 2.0
SLOW_PAGE_WARNING_SECONDS = 60
HEARTBEAT_SECONDS = 10
# 单页硬超时兜底（秒）：只针对真正卡死的页面；正常慢页不会误伤。
# 不设置页面硬超时。扫描件偶有极慢页，宁可持续显示心跳，也不丢弃该页。
PAGE_TIMEOUT_SECONDS: float | None = None
MEMORY_FALLBACK_DPI = (300, 250)
# Paddle 阶段每个独立子进程处理的文件数：子进程退出后操作系统强制回收其
# C++ 内存池（进程内 del/gc 无法释放）。批次越小越省内存、但重启开销越大。
PADDLE_BATCH_SIZE = 8
# 续跑状态目录名（相对 output_base），用于崩溃后断点续跑。
STATE_DIR_NAME = ".cma_ocr_resume"

CONTEXT_CHARS = 80
FUZZY_MATCHING = True
FUZZY_MIN_KEYWORD_LENGTH = 4
FUZZY_THRESHOLD = 0.84
# 工具最终定位到“文件+页码+关键词”，不是穷举页面里的每个相似字串。
# 每页每关键词最多保留少量低可信候选，避免长页产生大量近似窗口。
MAX_UNCERTAIN_HITS_PER_KEYWORD_PAGE = 3
# 仅用于短英文/数字关键词（如 CMA）的低风险 OCR 混淆折叠。它不会直接进入
# 主结果表：单引擎容错命中会进入候选表，二次精查或另一模型确认后才升为可信结果。
OCR_CONFUSION_FOLDING = str.maketrans({
    "0": "o", "o": "o",
    "1": "i", "i": "i", "l": "i",
    "2": "z", "z": "z",
    "4": "a", "a": "a",
    "5": "s", "s": "s",
    "6": "g", "g": "g",
    "8": "b", "b": "b",
})
SUSPECT_RECHECK_ENABLED = True
SUSPECT_RECHECK_DPI = 500
UNCERTAIN_MATCH_TYPES = frozenset({"fuzzy", "ocr_confusion"})
DEDUP_CONTEXT_THRESHOLD = 0.72
OUTPUT_PREFIX = "cma_results"
# ===========================================================================


@dataclass
class Hit:
    filename: str
    page: int
    keyword: str
    matched_text: str
    context: str
    engine: str
    match_type: str
    match_score: float
    start: int
    end: int
    text_origin: str = "ocr"


@dataclass
class MergedHit:
    filename: str
    page: int
    keyword: str
    hits: list[Hit] = field(default_factory=list)

    @property
    def engines(self) -> list[str]:
        return sorted({hit.engine for hit in self.hits})

    @property
    def agreement(self) -> str:
        return "双模型命中" if len(self.engines) > 1 else "单模型命中"

    @property
    def has_exact(self) -> bool:
        return any(hit.match_type == "exact_normalized" for hit in self.hits)

    @property
    def has_fuzzy(self) -> bool:
        return any(hit.match_type == "fuzzy" for hit in self.hits)

    @property
    def has_uncertain(self) -> bool:
        return any(hit.match_type in UNCERTAIN_MATCH_TYPES for hit in self.hits)

    @property
    def is_dual_engine(self) -> bool:
        return len(self.engines) > 1

    @property
    def confidence(self) -> str:
        # 置信度分级：L1 双引擎精确 > L2 双引擎一致 > L3 单引擎精确 > L4 单引擎候选。
        # 只看匹配质量与引擎一致性，不偏向任一引擎（不分主次）。
        if self.is_dual_engine:
            if self.has_exact and not self.has_fuzzy:
                return "L1-双引擎精确"
            return "L2-双引擎一致"
        if self.has_exact:
            return "L3-单引擎精确"
        return "L4-单引擎待复核"

    @property
    def is_candidate(self) -> bool:
        """单引擎且没有精确命中 → 待人工复核，不混入主结果表。"""
        return not self.is_dual_engine and self.has_uncertain and not self.has_exact

    def context_for(self, engine: str) -> str:
        candidates = [hit for hit in self.hits if hit.engine == engine]
        if not candidates:
            return ""
        return select_neutral_representative(candidates).context


@dataclass
class PageError:
    timestamp: str
    engine: str
    filename: str
    page: int | str
    attempt: int | str
    stage: str
    message: str


def configure_logging(output_dir: Path, latest_output_dir: Path | None = None) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    # 控制台只显示关键信息（文件进度、心跳、错误、汇总），逐页细节不刷屏；
    # 完整逐页日志仍写入 .log 文件，排查问题时查看日志即可。
    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    file_handler = logging.FileHandler(
        output_dir / f"{OUTPUT_PREFIX}.log", encoding="utf-8"
    )
    file_handler.setLevel(logging.DEBUG)
    handlers: list[logging.Handler] = [stream_handler, file_handler]
    if latest_output_dir is not None and latest_output_dir != output_dir:
        latest_output_dir.mkdir(parents=True, exist_ok=True)
        latest_handler = logging.FileHandler(
            latest_output_dir / f"{OUTPUT_PREFIX}.log", mode="w", encoding="utf-8"
        )
        latest_handler.setLevel(logging.DEBUG)
        handlers.append(latest_handler)
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
        handlers=handlers,
        force=True,
    )


def configure_worker_logging(log_path: Path) -> None:
    """子进程工作模式的日志：控制台只看关键信息，完整日志追加到主日志文件。"""
    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    logging.basicConfig(
        level=logging.DEBUG,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
        handlers=[stream_handler, file_handler],
        force=True,
    )


def configure_tesseract() -> None:
    for candidate in TESSERACT_EXE_CANDIDATES:
        if candidate.is_file():
            pytesseract.pytesseract.tesseract_cmd = str(candidate)
            return
    discovered = shutil.which("tesseract")
    if discovered:
        pytesseract.pytesseract.tesseract_cmd = discovered
        return
    searched = "；".join(str(c) for c in TESSERACT_EXE_CANDIDATES)
    raise FileNotFoundError(
        f"未找到 Tesseract。已尝试：{searched}；以及系统 PATH。"
        "请确认已安装并勾选中文语言包。"
    )


def init_paddle() -> Any | None:
    """Initialize the user's PaddleOCR 2.7.x stack for its independent run."""
    try:
        from paddleocr import PaddleOCR

        paddle_version = metadata.version("paddlepaddle")
        paddleocr_version = metadata.version("paddleocr")
        logging.info(
            "Paddle 环境：paddlepaddle=%s, paddleocr=%s",
            paddle_version,
            paddleocr_version,
        )
        if paddle_version != "2.6.2" or paddleocr_version != "2.7.0.3":
            logging.warning(
                "当前版本与代码目标版本不同；建议使用 "
                "paddlepaddle==2.6.2、paddleocr==2.7.0.3"
            )

        # PaddleOCR 2.7 uses use_angle_cls/cls, not the 3.x
        # use_textline_orientation/predict interface.
        return PaddleOCR(
            lang=PADDLE_LANG,
            use_angle_cls=True,
            use_gpu=False,
            enable_mkldnn=False,
            rec_batch_num=1,
            show_log=False,
        )
    except Exception as exc:
        logging.exception("PaddleOCR 初始化失败；Tesseract 结果仍会保留：%s", exc)
        return None


def render_page(page: fitz.Page, dpi: int) -> Image.Image:
    """Render without JPEG compression; the returned pixels are lossless RGB."""
    scale = dpi / 72.0
    pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
    try:
        image = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        image.info["render_dpi"] = dpi
        return image
    finally:
        del pix


def _print_progress(
    file_index: int,
    total_files: int,
    page_number: int,
    total_pages: int,
    filename: str,
) -> None:
    """在终端输出单行进度条，随每页完成而刷新（仅交互式终端显示）。

    与心跳日志配合：进度条反映"已推进到第几页"，心跳反映"当前页仍在识别"，
    二者共同用于区分"正在工作"与"卡死"。重定向输出时静默跳过，避免污染文件。
    """
    if not sys.stdout.isatty():
        return
    fraction = page_number / total_pages if total_pages else 1.0
    width = 16
    filled = int(round(fraction * width))
    bar = "#" * filled + "-" * (width - filled)
    short_name = filename if len(filename) <= 24 else filename[:21] + "..."
    line = (
        f"\r[文件 {file_index}/{total_files}] {short_name} "
        f"第{page_number}/{total_pages}页 |{bar}| {fraction * 100:5.1f}%"
    )
    # 行尾补空格，覆盖上一行更长的残留字符
    sys.stdout.write(line + " " * 8)
    sys.stdout.flush()


def _finish_progress() -> None:
    """结束当前进度条所在行，避免后续日志与进度条挤在同一行。"""
    if sys.stdout.isatty():
        sys.stdout.write("\n")
        sys.stdout.flush()


def _binarize_otsu(gray: Image.Image) -> Image.Image:
    """Otsu 全局二值化，用于扫描件/浅字提升 Tesseract 识别率。"""
    arr = np.asarray(gray)
    hist, _ = np.histogram(arr.ravel(), bins=256, range=(0, 256))
    total = arr.size
    sum_total = float(np.dot(np.arange(256), hist))
    weight_b = 0
    sum_b = 0.0
    best_thresh = 0
    best_between = 0.0
    for t in range(256):
        weight_b += int(hist[t])
        if weight_b == 0:
            continue
        weight_f = total - weight_b
        if weight_f == 0:
            break
        sum_b += t * hist[t]
        mean_b = sum_b / weight_b
        mean_f = (sum_total - sum_b) / weight_f
        between = weight_b * weight_f * (mean_b - mean_f) ** 2
        if between > best_between:
            best_between = between
            best_thresh = t
    binary = np.where(arr > best_thresh, 255, 0).astype(np.uint8)
    return Image.fromarray(binary, mode="L")


def prepare_for_tesseract(image: Image.Image, force_binarize: bool = False) -> Image.Image:
    gray = ImageOps.grayscale(image)
    gray = ImageOps.autocontrast(gray, cutoff=1)
    if BINARIZE or force_binarize:
        return _binarize_otsu(gray)
    return gray.filter(ImageFilter.UnsharpMask(radius=1.2, percent=130, threshold=3))


def ocr_tesseract(image: Image.Image, _engine: Any = None) -> str:
    prepared = prepare_for_tesseract(image)
    try:
        # 显式告知当前页的实际 DPI（内存降级页不会误报 400）。
        dpi = int(image.info.get("render_dpi", RENDER_DPI))
        config = f"--dpi {dpi} " + TESS_CONFIG
        return pytesseract.image_to_string(
            prepared,
            lang=TESS_LANG,
            config=config,
        ).strip()
    finally:
        prepared.close()


def ocr_tesseract_recheck(image: Image.Image, _engine: Any = None) -> str:
    """疑似页仅做一次更强的二值化复核，原始 OCR 结果仍会完整保留。"""
    prepared = prepare_for_tesseract(image, force_binarize=True)
    try:
        dpi = int(image.info.get("render_dpi", SUSPECT_RECHECK_DPI))
        config = f"--dpi {dpi} " + TESS_CONFIG
        return pytesseract.image_to_string(prepared, lang=TESS_LANG, config=config).strip()
    finally:
        prepared.close()


def _collect_rec_texts(value: Any, seen: set[int] | None = None) -> list[str]:
    """Read rec_texts from common PaddleOCR 2.x/3.x result shapes."""
    if value is None:
        return []
    if seen is None:
        seen = set()
    value_id = id(value)
    if value_id in seen:
        return []
    seen.add(value_id)

    if isinstance(value, dict):
        if "rec_texts" in value:
            texts = value["rec_texts"]
            if isinstance(texts, (list, tuple)):
                return [str(item) for item in texts if str(item).strip()]
            return [str(texts)] if str(texts).strip() else []
        collected: list[str] = []
        for nested in value.values():
            collected.extend(_collect_rec_texts(nested, seen))
        return collected

    if isinstance(value, (list, tuple)):
        collected = []
        for nested in value:
            collected.extend(_collect_rec_texts(nested, seen))
        return collected

    for attribute in ("rec_texts", "res", "json"):
        if hasattr(value, attribute):
            try:
                nested = getattr(value, attribute)
                nested = nested() if callable(nested) else nested
                found = _collect_rec_texts(nested, seen)
                if found:
                    return found
            except Exception:
                continue
    return []


def _collect_legacy_paddle_lines(result: Any) -> list[str]:
    lines: list[str] = []
    pages = result if isinstance(result, list) else [result]
    for page_result in pages:
        if not isinstance(page_result, list):
            continue
        for line in page_result:
            try:
                text = line[1][0]
                if isinstance(text, str) and text.strip():
                    lines.append(text)
            except (IndexError, TypeError):
                continue
    return lines


def ocr_paddle(image: Image.Image, engine: Any) -> str:
    # Paddle commonly expects an OpenCV-style BGR ndarray.
    rgb = np.asarray(image)
    bgr = np.ascontiguousarray(rgb[:, :, ::-1])
    result = None
    texts: list[str] = []
    try:
        # Deliberately use the PaddleOCR 2.7.0.3 API selected by the user.
        result = engine.ocr(bgr, cls=True)
        texts = _collect_legacy_paddle_lines(result) or _collect_rec_texts(result)
        return "\n".join(texts).strip()
    finally:
        del texts, result, bgr, rgb


def ocr_paddle_recheck(image: Image.Image, engine: Any) -> str:
    """Paddle 疑似页复核：保留灰度细节、拉开反差并轻微锐化后再识别。"""
    prepared = ImageOps.grayscale(image)
    enhanced: Image.Image | None = None
    try:
        enhanced = ImageOps.autocontrast(prepared, cutoff=1).filter(
            ImageFilter.UnsharpMask(radius=1.5, percent=150, threshold=2)
        ).convert("RGB")
        return ocr_paddle(enhanced, engine)
    finally:
        prepared.close()
        if enhanced is not None:
            enhanced.close()


class PageTimeoutError(Exception):
    """单页 OCR 超过兜底上限仍未返回，用于放弃真正卡死的页面。"""


def run_ocr_with_heartbeat(
    ocr_function: Any,
    image: Image.Image,
    engine: Any,
    engine_name: str,
    filename: str,
    page_number: int,
    page_timeout: float | None = PAGE_TIMEOUT_SECONDS,
) -> tuple[str, float]:
    """在守护线程中运行 OCR，并持续输出心跳。

    默认不设页面硬超时：扫描件的慢页会继续等待，不因时长而被放弃。保留可选
    参数只为将来排障时按需启用，正常运行不使用它。
    """
    result_queue: queue.Queue[tuple[str, Any]] = queue.Queue(maxsize=1)

    def _worker() -> None:
        try:
            result_queue.put(("ok", ocr_function(image, engine)))
        except Exception as exc:  # noqa: BLE001 - 把 OCR 异常原样回传给主线程
            result_queue.put(("err", exc))

    started = time.perf_counter()
    heartbeat = 0
    worker = threading.Thread(
        target=_worker, name=f"ocr-{engine_name}", daemon=True,
    )
    worker.start()

    while True:
        try:
            status, payload = result_queue.get(timeout=HEARTBEAT_SECONDS)
        except queue.Empty:
            elapsed = time.perf_counter() - started
            if page_timeout is not None and page_timeout > 0 and elapsed >= page_timeout:
                logging.error(
                    "[页面超时] 引擎=%s | 文件=%s | 页码=%d | 已耗时=%.0fs 超过 %.0fs，放弃该页",
                    engine_name, filename, page_number, elapsed, page_timeout,
                )
                raise PageTimeoutError(
                    f"{engine_name} | {filename} | 第{page_number}页识别超过"
                    f"{page_timeout:.0f}s，已放弃（结果若缺失需人工补）"
                )
            heartbeat += 1
            logging.info(
                "[心跳#%d 仍在识别] 引擎=%s | 文件=%s | 页码=%d | 已耗时=%.0fs",
                heartbeat, engine_name, filename, page_number, elapsed,
            )
            continue

        if status == "ok":
            return payload, time.perf_counter() - started
        # status == "err"：把 OCR 内部异常抛给上层重试逻辑。
        raise payload


def usable_character_count(text: str) -> int:
    return sum(char.isalnum() or "\u4e00" <= char <= "\u9fff" for char in text)


def is_memory_error(exc: BaseException) -> bool:
    message = f"{type(exc).__name__}: {exc}".casefold()
    markers = (
        "memoryerror", "out of memory", "bad allocation", "std::bad_alloc",
        "cannot allocate", "failed to allocate", "malloc", "create primitive",
    )
    return any(marker in message for marker in markers)


def render_dpi_candidates(primary_dpi: int) -> list[int]:
    """Keep 400 DPI first; lower DPI is used only after a memory failure."""
    candidates = [primary_dpi]
    candidates.extend(dpi for dpi in MEMORY_FALLBACK_DPI if dpi < primary_dpi)
    return list(dict.fromkeys(candidates))


def normalize_with_map(text: str) -> tuple[str, list[int]]:
    """Create a comparison string while retaining indices into original text."""
    normalized_chars: list[str] = []
    original_indices: list[int] = []
    for index, char in enumerate(text):
        expanded = unicodedata.normalize("NFKC", char).casefold()
        for normalized_char in expanded:
            if normalized_char.isalnum() or "\u4e00" <= normalized_char <= "\u9fff":
                normalized_chars.append(normalized_char)
                original_indices.append(index)
    return "".join(normalized_chars), original_indices


def normalize_for_comparison(text: str) -> str:
    return normalize_with_map(text)[0]


def _context(text: str, start: int, end: int) -> str:
    left = max(0, start - CONTEXT_CHARS)
    right = min(len(text), end + CONTEXT_CHARS)
    return re.sub(r"\s+", " ", text[left:right]).strip()


def _overlaps(start: int, end: int, spans: Iterable[tuple[int, int]]) -> bool:
    return any(start < old_end and end > old_start for old_start, old_end in spans)


def _is_ascii_token_for_confusion_matching(compact_keyword: str) -> bool:
    """只对 CMA 一类短英文/数字代号启用字符混淆，避免扩大中文词误报。"""
    return (
        3 <= len(compact_keyword) <= 8
        and compact_keyword.isascii()
        and compact_keyword.isalnum()
        and any(char.isalpha() for char in compact_keyword)
    )


def _fold_ocr_confusions(compact_text: str) -> str:
    return compact_text.translate(OCR_CONFUSION_FOLDING)


def find_keyword_hits(
    text: str,
    keywords: list[str],
    filename: str,
    page: int,
    engine: str,
    text_origin: str,
) -> list[Hit]:
    compact_text, index_map = normalize_with_map(text)
    if not compact_text or not index_map:
        return []

    hits: list[Hit] = []
    for keyword in keywords:
        compact_keyword = normalize_for_comparison(keyword)
        if not compact_keyword:
            continue

        accepted_spans: list[tuple[int, int]] = []
        search_from = 0
        while True:
            position = compact_text.find(compact_keyword, search_from)
            if position < 0:
                break
            original_start = index_map[position]
            original_end = index_map[position + len(compact_keyword) - 1] + 1
            accepted_spans.append((original_start, original_end))
            hits.append(
                Hit(
                    filename=filename,
                    page=page,
                    keyword=keyword,
                    matched_text=text[original_start:original_end],
                    context=_context(text, original_start, original_end),
                    engine=engine,
                    match_type="exact_normalized",
                    match_score=1.0,
                    start=original_start,
                    end=original_end,
                    text_origin=text_origin,
                )
            )
            search_from = position + max(1, len(compact_keyword))

        # 同一页、同一关键词已经出现过精确命中时，低可信的容错/模糊窗口
        # 不再提供额外定位价值，只会膨胀 raw 表并触发无必要的二次精查。
        if accepted_spans:
            continue

        # 短英文代号经常发生 A/4、O/0、S/5 这类单字符混淆。这里的命中
        # 标成低可信“容错”，绝不伪装为精确命中；后续二次精查和双模型合并
        # 才负责把它升级为可信结果。
        if _is_ascii_token_for_confusion_matching(compact_keyword):
            folded_keyword = _fold_ocr_confusions(compact_keyword)
            folded_text = _fold_ocr_confusions(compact_text)
            search_from = 0
            confusion_hits = 0
            while True:
                position = folded_text.find(folded_keyword, search_from)
                if position < 0:
                    break
                original_start = index_map[position]
                original_end = index_map[position + len(compact_keyword) - 1] + 1
                search_from = position + max(1, len(compact_keyword))
                if _overlaps(original_start, original_end, accepted_spans):
                    continue
                accepted_spans.append((original_start, original_end))
                hits.append(
                    Hit(
                        filename=filename,
                        page=page,
                        keyword=keyword,
                        matched_text=text[original_start:original_end],
                        context=_context(text, original_start, original_end),
                        engine=engine,
                        match_type="ocr_confusion",
                        match_score=0.92,
                        start=original_start,
                        end=original_end,
                        text_origin=text_origin,
                    )
                )
                confusion_hits += 1
                if confusion_hits >= MAX_UNCERTAIN_HITS_PER_KEYWORD_PAGE:
                    break

            # 字形混淆比通用模糊匹配更可解释；已有这类候选时不再叠加模糊窗口。
            if confusion_hits:
                continue

        if not FUZZY_MATCHING or len(compact_keyword) < FUZZY_MIN_KEYWORD_LENGTH:
            continue

        # OCR insertion/deletion errors are handled by checking nearby window sizes.
        candidate_windows: list[tuple[float, int, int]] = []
        key_length = len(compact_keyword)
        for window_length in range(max(2, key_length - 1), key_length + 2):
            for start in range(0, len(compact_text) - window_length + 1):
                end = start + window_length
                original_start = index_map[start]
                original_end = index_map[end - 1] + 1
                if _overlaps(original_start, original_end, accepted_spans):
                    continue
                matcher = SequenceMatcher(
                    None, compact_keyword, compact_text[start:end], autojunk=False
                )
                # quick_ratio 是 ratio 的上界：先粗筛可省去大量昂贵的完整 ratio 计算，
                # 且不改变最终结果（低于阈值的窗口无论如何也不会命中）。
                if matcher.quick_ratio() < FUZZY_THRESHOLD:
                    continue
                score = matcher.ratio()
                if score >= FUZZY_THRESHOLD:
                    candidate_windows.append((score, start, end))

        # Prefer the strongest non-overlapping fuzzy candidates.
        candidate_windows.sort(key=lambda item: (-item[0], item[1], item[2]))
        accepted_fuzzy = 0
        for score, start, end in candidate_windows:
            original_start = index_map[start]
            original_end = index_map[end - 1] + 1
            if _overlaps(original_start, original_end, accepted_spans):
                continue
            accepted_spans.append((original_start, original_end))
            hits.append(
                Hit(
                    filename=filename,
                    page=page,
                    keyword=keyword,
                    matched_text=text[original_start:original_end],
                    context=_context(text, original_start, original_end),
                    engine=engine,
                    match_type="fuzzy",
                    match_score=round(score, 4),
                    start=original_start,
                    end=original_end,
                    text_origin=text_origin,
                )
            )
            accepted_fuzzy += 1
            if accepted_fuzzy >= MAX_UNCERTAIN_HITS_PER_KEYWORD_PAGE:
                break
    return hits


def _confirmed_recheck_hits(
    engine_name: str,
    engine: Any,
    page: fitz.Page,
    relative_name: str,
    page_number: int,
    keywords: list[str],
    initial_hits: list[Hit],
    errors: list[PageError],
) -> list[Hit]:
    """对本引擎自己的低可信命中做一次高分辨率复核。

    触发条件、渲染、OCR 和结果判断都在同一引擎阶段内完成，不会读取另一模型
    的文本或命中，因此双模型仍是两个独立事件。只有复核得到精确命中才替换原
    低可信命中；否则原候选保留以供人工回查。
    """
    if not SUSPECT_RECHECK_ENABLED or not any(
        hit.match_type in UNCERTAIN_MATCH_TYPES for hit in initial_hits
    ):
        return initial_hits

    recheck_function = (
        ocr_tesseract_recheck if engine_name == "tesseract" else ocr_paddle_recheck
    )
    image: Image.Image | None = None
    try:
        logging.info(
            "[疑似页二次精查] 引擎=%s | 文件=%s | 页码=%d | DPI=%d",
            engine_name, relative_name, page_number, SUSPECT_RECHECK_DPI,
        )
        image = render_page(page, SUSPECT_RECHECK_DPI)
        recheck_text, elapsed = run_ocr_with_heartbeat(
            recheck_function,
            image,
            engine,
            f"{engine_name}-复核",
            relative_name,
            page_number,
        )
        recheck_hits = find_keyword_hits(
            recheck_text,
            keywords,
            relative_name,
            page_number,
            engine_name,
            "ocr_recheck",
        )
        exact_keywords = {
            normalize_for_comparison(hit.keyword)
            for hit in recheck_hits
            if hit.match_type == "exact_normalized"
        }
        if not exact_keywords:
            logging.info(
                "[疑似页二次精查完成] 引擎=%s | 文件=%s | 页码=%d | 耗时=%.1fs | 未新增精确命中",
                engine_name, relative_name, page_number, elapsed,
            )
            return initial_hits

        retained = [
            hit for hit in initial_hits
            if not (
                hit.match_type in UNCERTAIN_MATCH_TYPES
                and normalize_for_comparison(hit.keyword) in exact_keywords
            )
        ]
        confirmed = [
            hit for hit in recheck_hits if hit.match_type == "exact_normalized"
        ]
        logging.info(
            "[疑似页二次精查确认] 引擎=%s | 文件=%s | 页码=%d | 耗时=%.1fs | 精确命中=%d",
            engine_name, relative_name, page_number, elapsed, len(confirmed),
        )
        return retained + confirmed
    except Exception as exc:  # noqa: BLE001 - 复核失败不丢弃首次识别结果
        errors.append(
            PageError(
                timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                engine=engine_name,
                filename=relative_name,
                page=page_number,
                attempt="-",
                stage="疑似页二次精查",
                message=f"{type(exc).__name__}: {exc}",
            )
        )
        logging.warning(
            "[疑似页二次精查失败，保留首次结果] 引擎=%s | 文件=%s | 页码=%d | %s",
            engine_name, relative_name, page_number, exc,
        )
        return initial_hits
    finally:
        if image is not None:
            image.close()
        del image
        gc.collect()


def process_page_for_engine(
    engine_name: str,
    engine: Any,
    page: fitz.Page,
    relative_name: str,
    page_number: int,
    total_pages: int,
    file_index: int,
    total_files: int,
    keywords: list[str],
    errors: list[PageError],
) -> tuple[list[Hit], str, int]:
    """对单页、单引擎完成 OCR（含重试/PDF兜底/命中提取），返回该页命中。

    两个引擎各自独立调用本函数，互不读取对方结果，无主次之分。
    """
    ocr_function = ocr_tesseract if engine_name == "tesseract" else ocr_paddle
    page_text = ""
    text_origin = "ocr"
    page_error = False
    used_dpi = RENDER_DPI

    logging.debug(
        "[开始页面] 引擎=%s | 文件=%s | 文件进度=%d/%d | 页码=%d/%d",
        engine_name,
        relative_name,
        file_index,
        total_files,
        page_number,
        total_pages,
    )
    if engine_name == "paddle":
        _report_paddle_worker_stage(
            "正在识别页面",
            filename=relative_name,
            batch_file_index=file_index,
            batch_file_total=total_files,
            page=page_number,
            page_total=total_pages,
        )

    page_timed_out = False
    for dpi in render_dpi_candidates(RENDER_DPI):
        used_dpi = dpi
        move_to_lower_dpi = False
        for attempt in range(1, MAX_RETRIES + 1):
            image: Image.Image | None = None
            try:
                image = render_page(page, dpi)
                page_text, elapsed = run_ocr_with_heartbeat(
                    ocr_function, image, engine, engine_name, relative_name, page_number,
                )
                if elapsed > SLOW_PAGE_WARNING_SECONDS:
                    logging.warning(
                        "%s | %s | 第%d页耗时 %.1fs | DPI=%d",
                        engine_name, relative_name, page_number, elapsed, dpi,
                    )
                break
            except PageTimeoutError as exc:
                page_error = True
                page_timed_out = True
                errors.append(
                    PageError(
                        timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                        engine=engine_name,
                        filename=relative_name,
                        page=page_number,
                        attempt=attempt,
                        stage="页面超时",
                        message=str(exc),
                    )
                )
                break
            except Exception as exc:
                page_error = True
                memory_failure = is_memory_error(exc)
                errors.append(
                    PageError(
                        timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                        engine=engine_name,
                        filename=relative_name,
                        page=page_number,
                        attempt=attempt,
                        stage="内存降级重试" if memory_failure else "OCR识别",
                        message=f"DPI={dpi} | {type(exc).__name__}: {exc}",
                    )
                )
                if memory_failure:
                    logging.warning(
                        "[内存不足] 引擎=%s | 文件=%s | 页码=%d | DPI=%d；"
                        "将仅降低当前页 DPI 重试",
                        engine_name, relative_name, page_number, dpi,
                    )
                    move_to_lower_dpi = True
                    break
                logging.warning(
                    "%s | %s | 第%d页 | DPI=%d | 尝试%d/%d失败：%s",
                    engine_name, relative_name, page_number, dpi,
                    attempt, MAX_RETRIES, exc,
                )
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY_SECONDS)
            finally:
                if image is not None:
                    image.close()
                del image
                gc.collect()
        if page_timed_out:
            break
        if page_text:
            if dpi != RENDER_DPI:
                logging.info(
                    "[内存降级成功] 引擎=%s | 文件=%s | 页码=%d | DPI=%d",
                    engine_name, relative_name, page_number, dpi,
                )
            break
        if not move_to_lower_dpi:
            break

    if (
        usable_character_count(page_text) < MIN_OCR_TEXT_LENGTH
        and USE_PDF_TEXT_AS_FALLBACK
    ):
        fallback = page.get_text("text").strip()
        if usable_character_count(fallback) > usable_character_count(page_text):
            page_text = fallback
            text_origin = "pdf_text_fallback"
            logging.info(
                "%s | %s | 第%d页使用PDF文本层兜底",
                engine_name,
                relative_name,
                page_number,
            )

    page_hits = find_keyword_hits(
        page_text,
        keywords,
        relative_name,
        page_number,
        engine_name,
        text_origin,
    )
    page_hits = _confirmed_recheck_hits(
        engine_name,
        engine,
        page,
        relative_name,
        page_number,
        keywords,
        page_hits,
        errors,
    )

    if not page_text:
        errors.append(
            PageError(
                timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                engine=engine_name,
                filename=relative_name,
                page=page_number,
                attempt=MAX_RETRIES if page_error else "-",
                stage="页面结果",
                message=f"OCR及PDF文本层均未获得文本（文本来源={text_origin}）",
            )
        )
        logging.warning(
            "[页面无结果] 引擎=%s | 文件=%s | 页码=%d",
            engine_name,
            relative_name,
            page_number,
        )

    logging.debug(
        "[完成页面] 引擎=%s | 文件=%s | 页码=%d/%d | 命中=%d | 文本来源=%s",
        engine_name,
        relative_name,
        page_number,
        total_pages,
        len(page_hits),
        text_origin,
    )
    if engine_name == "paddle":
        _report_paddle_worker_stage(
            "页面识别完成",
            filename=relative_name,
            batch_file_index=file_index,
            batch_file_total=total_files,
            page=page_number,
            page_total=total_pages,
            hits=len(page_hits),
        )
    status = "text" if page_text else ("error" if page_error else "empty")
    return page_hits, status, used_dpi


def run_single_engine(
    engine_name: str,
    engine: Any,
    pdf_files: list[tuple[Path, str]],
    keywords: list[str],
    errors: list[PageError],
    raw_path: Path,
    state: dict,
    state_dir: Path,
) -> tuple[list[Hit], dict[str, int]]:
    """Run one complete engine pass; it shares no page image with the other."""
    hits: list[Hit] = []
    stats = {
        "pages": 0, "text": 0, "empty": 0, "error": 0, "fallback": 0,
        "skipped_files": 0,
    }
    done = set(state.get(f"{engine_name}_done", []))
    logging.info("========== %s 独立识别阶段开始 ==========", engine_name)

    for file_index, (pdf_path, relative_name) in enumerate(pdf_files, 1):
        if relative_name in done:
            stats["skipped_files"] += 1
            logging.info("[%d/%d] 跳过（已完成） %s", file_index, len(pdf_files), relative_name)
            continue

        file_hits: list[Hit] = []
        file_failed = False
        logging.info("[%d/%d] %s", file_index, len(pdf_files), relative_name)
        try:
            with fitz.open(pdf_path) as document:
                total_pages = len(document)
                for page_index, page in enumerate(document):
                    page_number = page_index + 1
                    try:
                        page_hits, status, used_dpi = process_page_for_engine(
                            engine_name, engine, page, relative_name, page_number,
                            total_pages, file_index, len(pdf_files), keywords, errors,
                        )
                        hits.extend(page_hits)
                        file_hits.extend(page_hits)
                        stats["pages"] += 1
                        stats[status] += 1
                        if used_dpi != RENDER_DPI:
                            stats["fallback"] += 1
                    except Exception as exc:
                        stats["pages"] += 1
                        stats["error"] += 1
                        errors.append(PageError(
                            timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                            engine=engine_name, filename=relative_name, page=page_number,
                            attempt="-", stage="单页保护",
                            message=f"{type(exc).__name__}: {exc}",
                        ))
                        logging.exception(
                            "[单页失败但继续] 引擎=%s | 文件=%s | 页码=%d",
                            engine_name, relative_name, page_number,
                        )
                    finally:
                        gc.collect()
                    _print_progress(
                        file_index, len(pdf_files),
                        page_number, total_pages, relative_name,
                    )
        except Exception as exc:
            file_failed = True
            errors.append(
                PageError(
                    timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                    engine=engine_name,
                    filename=relative_name,
                    page="-",
                    attempt="-",
                    stage="打开或遍历PDF",
                    message=f"文件 {relative_name} 无法处理：{type(exc).__name__}: {exc}",
                )
            )
            logging.exception("无法处理 %s：%s", pdf_path, exc)
        _finish_progress()
        gc.collect()

        if not file_failed:
            # 文件全部页处理完：立即落盘并标记完成，崩溃不丢该文件结果。
            append_raw_hits(raw_path, file_hits)
            done.add(relative_name)
            state[f"{engine_name}_done"] = sorted(done)
            _save_state(state_dir, state)

    logging.info(
        "========== %s 独立识别阶段完成 | 总页=%d | 有文本=%d | "
        "空页=%d | 失败=%d | 内存降级页=%d | 本轮跳过已完成文件=%d | 本次落盘命中=%d ==========",
        engine_name, stats["pages"], stats["text"], stats["empty"],
        stats["error"], stats["fallback"], stats["skipped_files"], len(hits),
    )
    return hits, stats


def _paddle_worker_main(batch_json_path: str) -> int:
    """子进程工作模式：处理一批 Paddle 文件后退出，让操作系统回收其 C++ 内存池。

    退出码约定：0=本批正常完成；2=进程内可捕获异常/批次文件损坏；3=Paddle 初始化失败。
    """
    try:
        spec = json.loads(Path(batch_json_path).read_text(encoding="utf-8"))
    except (ValueError, OSError) as exc:
        print(f"[子进程] 无法读取批次文件 {batch_json_path}: {exc}")
        return 2
    files = [(Path(p), rel) for p, rel in spec["files"]]
    keywords = spec["keywords"]
    raw_path = Path(spec["raw_path"])
    state_dir = Path(spec["state_dir"])
    log_path = Path(spec["log_path"])

    _report_paddle_worker_stage("已读取批次任务，正在配置日志")
    configure_worker_logging(log_path)
    logging.info("========== Paddle 子进程启动（本批 %d 个文件） ==========", len(files))

    _report_paddle_worker_stage("正在初始化 PaddleOCR 模型（模型缓存会复用用户目录）")
    engine = init_paddle()
    if engine is None:
        _report_paddle_worker_stage("PaddleOCR 初始化失败")
        logging.error("Paddle 子进程初始化失败，返回退出码 3")
        return 3
    # PaddleOCR initialization may reconfigure Python's root logger. Restore
    # our dedicated append-only worker log before page processing so the web
    # UI and later diagnostics retain every Paddle file/page event.
    configure_worker_logging(log_path)
    logging.info("PaddleOCR 模型初始化完成，进入本批页级识别")
    _report_paddle_worker_stage("PaddleOCR 初始化完成，开始识别本批文件")

    state = _load_state(state_dir)
    errors: list[PageError] = []
    try:
        hits, stats = run_single_engine(
            "paddle", engine, files, keywords, errors, raw_path, state, state_dir,
        )
    except BaseException as exc:  # noqa: BLE001 - OOM 大多直接杀进程，这里兜底可捕获异常
        logging.exception("Paddle 子进程异常退出：%s", exc)
        return 2

    (state_dir / "paddle_batch_stats.json").write_text(
        json.dumps(stats, ensure_ascii=False), encoding="utf-8"
    )
    (state_dir / "paddle_batch_errors.json").write_text(
        json.dumps([asdict(error) for error in errors], ensure_ascii=False),
        encoding="utf-8",
    )
    logging.info("========== Paddle 子进程完成，本批命中 %d ==========", len(hits))
    _report_paddle_worker_stage("本批识别完成，子进程即将退出")
    return 0


def _read_paddle_worker_stage(status_path: Path) -> str:
    try:
        payload = json.loads(status_path.read_text(encoding="utf-8"))
        stage = str(payload.get("stage", "")).strip()
        if payload.get("filename"):
            stage += (
                f" | 文件={payload['filename']}"
                f" | 页码={payload.get('page', '-')}/{payload.get('page_total', '-')}"
            )
        return stage or "子进程正在运行，尚未报告阶段"
    except (OSError, ValueError, TypeError):
        return "子进程正在启动，等待首个状态回报"


def _wait_for_paddle_worker(command: list[str], status_path: Path) -> int:
    """等待 Paddle 批处理子进程，同时给出启动/导入/模型初始化阶段心跳。

    不设总超时、不杀死进程。这样大型模型初次初始化或慢页识别会持续运行，而
    控制台仍能明确显示 PID 和最近阶段。
    """
    status_path.unlink(missing_ok=True)
    environment = os.environ.copy()
    environment["CMA_PADDLE_WORKER_STATUS_FILE"] = str(status_path)
    process = subprocess.Popen(command, env=environment)
    logging.info("Paddle 子进程已创建 | PID=%d", process.pid)
    heartbeat = 0
    while True:
        return_code = process.poll()
        if return_code is not None:
            return return_code
        time.sleep(HEARTBEAT_SECONDS)
        heartbeat += 1
        logging.info(
            "[Paddle 子进程心跳#%d] PID=%d | 阶段=%s",
            heartbeat,
            process.pid,
            _read_paddle_worker_stage(status_path),
        )


def run_paddle_isolated(
    pdf_files: list[tuple[Path, str]],
    keywords: list[str],
    errors: list[PageError],
    raw_path: Path,
    state_dir: Path,
    log_path: Path,
    batch_size: int = PADDLE_BATCH_SIZE,
) -> tuple[dict[str, int], bool]:
    """分批用独立子进程跑 Paddle：每个子进程处理一批后退出，操作系统强制回收其
    C++ 内存池。进程内 del/gc 无法释放该内存，这是唯一能根治长期运行 OOM 的办法。
    """
    state = _load_state(state_dir)
    aggregate = {"pages": 0, "text": 0, "empty": 0, "error": 0, "fallback": 0}
    paddle_ran = False
    current_batch = batch_size

    while True:
        done = set(state.get("paddle_done", []))
        remaining = [item for item in pdf_files if item[1] not in done]
        if not remaining:
            break
        batch = remaining[:current_batch]
        before = len(done)

        batch_file = state_dir / "paddle_batch.json"
        batch_file.write_text(
            json.dumps(
                {
                    "files": [[str(p), rel] for p, rel in batch],
                    "keywords": keywords,
                    "raw_path": str(raw_path),
                    "state_dir": str(state_dir),
                    "log_path": str(log_path),
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        logging.info(
            "启动 Paddle 子进程处理本批 %d 个文件（已完成 %d / 共 %d）",
            len(batch), before, len(pdf_files),
        )
        status_file = state_dir / "paddle_worker_status.json"
        return_code = _wait_for_paddle_worker(
            [
                sys.executable,
                str(Path(__file__).resolve()),
                "--_paddle_worker",
                str(batch_file),
            ],
            status_file,
        )
        batch_file.unlink(missing_ok=True)
        status_file.unlink(missing_ok=True)

        stats_file = state_dir / "paddle_batch_stats.json"
        if stats_file.is_file():
            try:
                batch_stats = json.loads(stats_file.read_text(encoding="utf-8"))
                for key in aggregate:
                    aggregate[key] += int(batch_stats.get(key, 0))
            except (ValueError, OSError):
                pass
            stats_file.unlink(missing_ok=True)

        errors_file = state_dir / "paddle_batch_errors.json"
        if errors_file.is_file():
            try:
                for item in json.loads(errors_file.read_text(encoding="utf-8")):
                    errors.append(PageError(**item))
            except (ValueError, OSError):
                pass
            errors_file.unlink(missing_ok=True)

        state = _load_state(state_dir)
        new_done = set(state.get("paddle_done", []))

        if return_code == 3:
            errors.append(
                PageError(
                    timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                    engine="paddle",
                    filename="-",
                    page="-",
                    attempt="-",
                    stage="引擎初始化",
                    message="PaddleOCR 初始化失败（子进程），跳过 Paddle 阶段",
                )
            )
            logging.error("PaddleOCR 初始化失败，跳过 Paddle 阶段")
            break

        paddle_ran = True
        if len(new_done) <= before:
            # 子进程连一个文件都没完成：大概率崩溃/OOM。缩小批次重试，
            # 单文件仍失败则记错跳过，避免死循环。
            if current_batch > 1:
                current_batch = max(1, current_batch // 2)
                logging.warning(
                    "Paddle 子进程本批未推进（疑似崩溃/OOM），批次缩小为 %d 后重试",
                    current_batch,
                )
            else:
                skipped = batch[0][1]
                new_done.add(skipped)
                state["paddle_done"] = sorted(new_done)
                _save_state(state_dir, state)
                errors.append(
                    PageError(
                        timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                        engine="paddle",
                        filename=skipped,
                        page="-",
                        attempt="-",
                        stage="单文件子进程失败",
                        message="该文件在独立子进程中反复失败，已跳过",
                    )
                )
                logging.error("Paddle 子进程反复失败，跳过文件：%s", skipped)
                current_batch = batch_size
        else:
            # 正常推进，恢复默认批次大小。
            current_batch = batch_size

    return aggregate, paddle_ran


def context_similarity(left: str, right: str) -> float:
    left_normalized = normalize_for_comparison(left)
    right_normalized = normalize_for_comparison(right)
    if not left_normalized or not right_normalized:
        return 0.0
    return SequenceMatcher(
        None, left_normalized, right_normalized, autojunk=False
    ).ratio()


def select_neutral_representative(hits: list[Hit]) -> Hit:
    # Engine name is deliberately absent from the ranking: neither model is primary.
    return sorted(
        hits,
        key=lambda hit: (
            -hit.match_score,
            -usable_character_count(hit.context),
            normalize_for_comparison(hit.context),
        ),
    )[0]


def merge_and_deduplicate(all_hits: list[Hit]) -> list[MergedHit]:
    groups: dict[tuple[str, int, str], list[MergedHit]] = {}
    ordered_hits = sorted(
        all_hits,
        key=lambda hit: (
            hit.filename.casefold(),
            hit.page,
            hit.keyword.casefold(),
            hit.start,
            hit.engine,
        ),
    )

    for hit in ordered_hits:
        key = (hit.filename, hit.page, normalize_for_comparison(hit.keyword))
        clusters = groups.setdefault(key, [])
        best_cluster: MergedHit | None = None
        best_similarity = -1.0
        for cluster in clusters:
            similarity = max(
                context_similarity(hit.context, old_hit.context)
                for old_hit in cluster.hits
            )
            if similarity >= DEDUP_CONTEXT_THRESHOLD and similarity > best_similarity:
                best_similarity = similarity
                best_cluster = cluster

        if best_cluster is None:
            clusters.append(
                MergedHit(
                    filename=hit.filename,
                    page=hit.page,
                    keyword=hit.keyword,
                    hits=[hit],
                )
            )
        else:
            best_cluster.hits.append(hit)

    merged = [cluster for clusters in groups.values() for cluster in clusters]
    return sorted(
        merged,
        key=lambda item: (item.filename.casefold(), item.page, item.keyword.casefold()),
    )


def excel_safe(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


RAW_CSV_HEADER = [
    "文件名",
    "页码",
    "关键词",
    "实际匹配文本",
    "匹配方式",
    "匹配得分",
    "上下文",
    "识别引擎",
    "文本来源",
]


def _hit_to_row(hit: Hit) -> list[Any]:
    return [
        excel_safe(hit.filename),
        hit.page,
        excel_safe(hit.keyword),
        excel_safe(hit.matched_text),
        hit.match_type,
        hit.match_score,
        excel_safe(hit.context),
        hit.engine,
        hit.text_origin,
    ]


def append_raw_hits(path: Path, hits: list[Hit]) -> None:
    """以追加方式落盘命中；空结果也确保表头存在，便于续跑读回。"""
    write_header = not path.exists() or path.stat().st_size == 0
    with path.open("a", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        if write_header:
            writer.writerow(RAW_CSV_HEADER)
        for hit in hits:
            writer.writerow(_hit_to_row(hit))


def read_raw_hits(path: Path) -> list[Hit]:
    """从 raw CSV 读回命中，用于续跑后的最终合并。

    raw CSV 是为断点续跑设计的轻量格式，故意没有保存 ``start/end``。
    它们只用于本轮内存中的排序，读回后给 0 即可；不能因缺少这两个字段
    把整行命中静默丢弃。
    """
    if not path.exists():
        logging.warning("原始命中表不存在：%s", path)
        return []
    hits: list[Hit] = []
    total_rows = 0
    rejected_rows = 0
    rejection_examples: list[str] = []
    with path.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        next(reader, None)  # 跳过表头
        for row_number, row in enumerate(reader, start=2):
            total_rows += 1
            if len(row) < 9:
                rejected_rows += 1
                if len(rejection_examples) < 3:
                    rejection_examples.append(f"第{row_number}行列数不足（{len(row)}列）")
                continue
            try:
                hits.append(
                    Hit(
                        filename=row[0],
                        page=int(row[1]),
                        keyword=row[2],
                        matched_text=row[3],
                        match_type=row[4],
                        match_score=float(row[5]),
                        context=row[6],
                        engine=row[7],
                        text_origin=row[8],
                        # start/end 没有写入 raw CSV；读回的命中只用于合并，
                        # 不再依赖这两个定位偏移量。
                        start=0,
                        end=0,
                    )
                )
            except (ValueError, TypeError) as exc:
                rejected_rows += 1
                if len(rejection_examples) < 3:
                    rejection_examples.append(f"第{row_number}行：{type(exc).__name__}: {exc}")
                continue
    logging.info(
        "原始命中表读回：%s | 数据行=%d | 成功=%d | 拒绝=%d",
        path.name, total_rows, len(hits), rejected_rows,
    )
    if rejection_examples:
        logging.warning("原始命中表被拒绝的样例：%s", "；".join(rejection_examples))
    return hits


def _state_dir(output_base: Path) -> Path:
    return output_base / STATE_DIR_NAME


def _load_state(state_dir: Path) -> dict:
    path = state_dir / "state.json"
    default = {
        "input_dir": None,
        "keywords": None,
        "dpi": None,
        "files": None,
        "tesseract_done": [],
        "paddle_done": [],
    }
    if not path.is_file():
        return default
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return default
    for key, value in default.items():
        data.setdefault(key, value)
    return data


def _save_state(state_dir: Path, state: dict) -> None:
    state_dir.mkdir(parents=True, exist_ok=True)
    (state_dir / "state.json").write_text(
        json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _clear_state(state_dir: Path) -> None:
    shutil.rmtree(state_dir, ignore_errors=True)


def write_raw_csv(path: Path, hits: list[Hit]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(RAW_CSV_HEADER)
        for hit in hits:
            writer.writerow(_hit_to_row(hit))


def write_merged_csv(path: Path, merged_hits: list[MergedHit]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                "文件名",
                "页码",
                "关键词",
                "置信度",
                "来源引擎",
                "代表上下文",
                "Tesseract上下文",
                "PaddleOCR上下文",
                "最高匹配得分",
                "原始命中数",
            ]
        )
        for item in merged_hits:
            representative = select_neutral_representative(item.hits)
            writer.writerow(
                [
                    excel_safe(item.filename),
                    item.page,
                    excel_safe(item.keyword),
                    item.confidence,
                    "+".join(item.engines),
                    excel_safe(representative.context),
                    excel_safe(item.context_for("tesseract")),
                    excel_safe(item.context_for("paddle")),
                    max(hit.match_score for hit in item.hits),
                    len(item.hits),
                ]
            )


def write_candidate_csv(path: Path, candidates: list[MergedHit]) -> None:
    """导出单引擎模糊命中的候选清单，供人工复核（不进入主结果表）。"""
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(
            [
                "文件名",
                "页码",
                "关键词",
                "实际匹配文本",
                "匹配方式",
                "匹配得分",
                "上下文",
                "识别引擎",
                "文本来源",
                "原始命中数",
            ]
        )
        for item in candidates:
            representative = select_neutral_representative(item.hits)
            writer.writerow(
                [
                    excel_safe(item.filename),
                    item.page,
                    excel_safe(item.keyword),
                    excel_safe(representative.matched_text),
                    representative.match_type,
                    representative.match_score,
                    excel_safe(representative.context),
                    "+".join(item.engines),
                    representative.text_origin,
                    len(item.hits),
                ]
            )


def write_error_csv(path: Path, errors: list[PageError]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["时间", "识别引擎", "文件名", "页码", "尝试次数", "阶段", "错误信息"])
        for error in errors:
            writer.writerow(
                [
                    error.timestamp,
                    error.engine,
                    excel_safe(error.filename),
                    error.page,
                    error.attempt,
                    error.stage,
                    excel_safe(error.message),
                ]
            )


def _append_sheet_header(sheet: Any, headers: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _finish_sheet_layout(sheet: Any) -> None:
    # 适度限制列宽，兼顾 Excel 可读性与大量上下文的浏览效率。
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        maximum = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(maximum + 2, 10), 60)
    sheet.auto_filter.ref = sheet.dimensions


def write_results_xlsx(
    path: Path,
    trusted_hits: list[MergedHit],
    candidate_hits: list[MergedHit],
    tesseract_hits: list[Hit],
    paddle_hits: list[Hit],
    errors: list[PageError],
) -> None:
    """导出与 CSV 对应的单一 XLSX 工作簿，便于筛选、汇总与人工回查。"""
    from openpyxl import Workbook

    workbook = Workbook()
    summary = workbook.active
    summary.title = "运行摘要"
    _append_sheet_header(summary, ["项目", "数值"])
    summary_rows = [
        ("可信命中（主表）", len(trusted_hits)),
        ("待复核候选", len(candidate_hits)),
        ("Tesseract 原始命中", len(tesseract_hits)),
        ("PaddleOCR 原始命中", len(paddle_hits)),
        ("错误/告警条数", len(errors)),
        ("说明", "L4 候选不会混入主结果；可按文件名、页码在原 PDF 中复查。"),
    ]
    for row in summary_rows:
        summary.append(row)
    _finish_sheet_layout(summary)

    merged_headers = [
        "文件名", "页码", "关键词", "置信度", "来源引擎", "代表上下文",
        "Tesseract上下文", "PaddleOCR上下文", "最高匹配得分", "原始命中数",
    ]
    merged_sheet = workbook.create_sheet("主结果")
    _append_sheet_header(merged_sheet, merged_headers)
    for item in trusted_hits:
        representative = select_neutral_representative(item.hits)
        merged_sheet.append([
            item.filename, item.page, item.keyword, item.confidence,
            "+".join(item.engines), representative.context,
            item.context_for("tesseract"), item.context_for("paddle"),
            max(hit.match_score for hit in item.hits), len(item.hits),
        ])
    _finish_sheet_layout(merged_sheet)

    candidate_sheet = workbook.create_sheet("待复核候选")
    candidate_headers = [
        "文件名", "页码", "关键词", "实际匹配文本", "匹配方式", "匹配得分",
        "上下文", "识别引擎", "文本来源", "原始命中数",
    ]
    _append_sheet_header(candidate_sheet, candidate_headers)
    for item in candidate_hits:
        representative = select_neutral_representative(item.hits)
        candidate_sheet.append([
            item.filename, item.page, item.keyword, representative.matched_text,
            representative.match_type, representative.match_score, representative.context,
            "+".join(item.engines), representative.text_origin, len(item.hits),
        ])
    _finish_sheet_layout(candidate_sheet)

    for title, raw_hits in (("Tesseract原始", tesseract_hits), ("Paddle原始", paddle_hits)):
        sheet = workbook.create_sheet(title)
        _append_sheet_header(sheet, RAW_CSV_HEADER)
        for hit in raw_hits:
            sheet.append(_hit_to_row(hit))
        _finish_sheet_layout(sheet)

    errors_sheet = workbook.create_sheet("错误日志")
    _append_sheet_header(errors_sheet, ["时间", "识别引擎", "文件名", "页码", "尝试次数", "阶段", "错误信息"])
    for error in errors:
        errors_sheet.append([
            error.timestamp, error.engine, error.filename, error.page,
            error.attempt, error.stage, error.message,
        ])
    _finish_sheet_layout(errors_sheet)
    workbook.save(path)
    workbook.close()


def rebuild_results_from_raw(output_base: Path) -> int:
    """Regenerate result tables without calling either OCR engine.

    This is intentionally non-destructive when both raw tables contain zero
    recoverable hits: the caller can inspect the resume files before choosing a
    full re-run.
    """
    state_dir = output_base / STATE_DIR_NAME
    tesseract_raw_path = state_dir / "tesseract_raw.csv"
    paddle_raw_path = state_dir / "paddle_raw.csv"
    if not tesseract_raw_path.exists() and not paddle_raw_path.exists():
        print(
            f"[无法重建] 未找到原始命中表：{state_dir}\n"
            "这通常表示该目录没有可恢复的断点数据。"
        )
        return 4

    try:
        run_lock = acquire_single_instance_lock(output_base)
    except RuntimeError as exc:
        print(f"[重复运行保护] {exc}")
        return 3
    _ = run_lock  # Keep the file lock alive until the process exits.

    run_name = time.strftime("rebuild_%Y%m%d_%H%M%S") + f"_{os.getpid()}"
    output_dir = output_base / "runs" / run_name
    configure_logging(output_dir, output_base)
    logging.info("========== 仅重建结果：不执行 OCR ========== ")
    logging.info("断点目录：%s", state_dir)

    tesseract_hits = read_raw_hits(tesseract_raw_path)
    paddle_hits = read_raw_hits(paddle_raw_path)
    if not tesseract_hits and not paddle_hits:
        logging.error(
            "两个原始命中表均没有可恢复的命中；为保护已有结果，本次没有覆盖任何结果 CSV。"
        )
        return 4

    merged_hits = merge_and_deduplicate(tesseract_hits + paddle_hits)
    trusted_hits = [item for item in merged_hits if not item.is_candidate]
    candidate_hits = [item for item in merged_hits if item.is_candidate]

    for target_dir in (output_dir, output_base):
        write_raw_csv(target_dir / f"{OUTPUT_PREFIX}_tesseract_raw.csv", tesseract_hits)
        write_raw_csv(target_dir / f"{OUTPUT_PREFIX}_paddle_raw.csv", paddle_hits)
        write_merged_csv(target_dir / f"{OUTPUT_PREFIX}_merged.csv", trusted_hits)
        write_candidate_csv(target_dir / f"{OUTPUT_PREFIX}_candidates.csv", candidate_hits)
        write_results_xlsx(
            target_dir / f"{OUTPUT_PREFIX}.xlsx",
            trusted_hits,
            candidate_hits,
            tesseract_hits,
            paddle_hits,
            [],
        )

    (output_dir / "rebuild_note.txt").write_text(
        "本目录由 --rebuild-results 从 .cma_ocr_resume 原始命中表重建；未重新执行 OCR。\n",
        encoding="utf-8",
    )
    (output_base / "CMA_OCR_最近一次结果.txt").write_text(
        str(output_dir), encoding="utf-8"
    )
    logging.info("Tesseract 读回命中：%d", len(tesseract_hits))
    logging.info("PaddleOCR 读回命中：%d", len(paddle_hits))
    logging.info("合并去重后命中：%d", len(merged_hits))
    logging.info("可信命中（主表）：%d", len(trusted_hits))
    logging.info("待复核候选（模糊）：%d", len(candidate_hits))
    logging.info("重建完成：%s", output_dir)
    return 0


def parse_keywords(raw: str) -> list[str]:
    raw = raw.strip()
    if not raw:
        return ["CMA"]

    if re.search(r"[,，、;；]", raw):
        parts = re.split(r"[,，、;；]+", raw)
    else:
        # Quoted phrases are retained: CMA "quality management system"
        parts = shlex.split(raw, posix=False)

    keywords: list[str] = []
    seen: set[str] = set()
    for part in parts:
        cleaned = part.strip().strip('"\'')
        identity = normalize_for_comparison(cleaned)
        if cleaned and identity and identity not in seen:
            seen.add(identity)
            keywords.append(cleaned)
    return keywords or ["CMA"]


def discover_pdfs(input_dir: Path) -> list[tuple[Path, str]]:
    files = sorted(
        input_dir.rglob("*.pdf"),
        key=lambda path: str(path.relative_to(input_dir)).casefold(),
    )
    return [(path, str(path.relative_to(input_dir))) for path in files]


def build_file_manifest(pdf_files: list[tuple[Path, str]]) -> list[dict[str, int | str]]:
    """Return a lightweight job fingerprint so a changed PDF cannot be skipped on resume."""
    manifest: list[dict[str, int | str]] = []
    for path, relative_name in pdf_files:
        stat = path.stat()
        manifest.append(
            {
                "filename": relative_name,
                "size": stat.st_size,
                "mtime_ns": stat.st_mtime_ns,
            }
        )
    return manifest


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="双引擎独立 OCR 与关键词提取")
    parser.add_argument("input_dir", nargs="?", help="PDF 根目录")
    parser.add_argument("-k", "--keywords", help="关键词；多个词用逗号分隔")
    parser.add_argument(
        "-o", "--output-dir",
        help="结果保存目录（不存在会自动新建）；不指定则运行后提示输入，直接回车默认当前目录",
    )
    parser.add_argument(
        "--dpi", type=int, default=RENDER_DPI,
        help=f"PDF 渲染分辨率，小字/密集排版可调高（默认 {RENDER_DPI}）",
    )
    parser.add_argument(
        "--binarize", action="store_true",
        help="Tesseract 预处理开启二值化（适合扫描件/浅字）",
    )
    parser.add_argument(
        "--rebuild-results", action="store_true",
        help="不重新 OCR；从 .cma_ocr_resume 的原始命中表重新生成结果",
    )
    parser.add_argument(
        "--check-env-only", action="store_true",
        help=argparse.SUPPRESS,
    )
    return parser


def clean_path_argument(raw_path: str) -> str:
    """Remove whitespace and stray shell quotes around a Windows path."""
    return raw_path.strip().strip('"').strip()


def acquire_single_instance_lock(output_base: Path) -> Any:
    """Prevent two accidental launches from competing for RAM on Windows."""
    import msvcrt

    output_base.mkdir(parents=True, exist_ok=True)
    lock_path = output_base / ".cma_ocr_running.lock"
    handle = lock_path.open("a+b")
    if handle.tell() == 0:
        handle.write(b"0")
        handle.flush()
    handle.seek(0)
    try:
        msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
    except OSError:
        handle.close()
        raise RuntimeError(
            "检测到另一个 CMA OCR 正在运行。为避免两个 Paddle 争抢内存，"
            "本次未启动；请等上一个窗口完成后再运行。"
        )

    def _release() -> None:
        try:
            handle.seek(0)
            msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
            handle.close()
        except Exception:
            pass

    atexit.register(_release)
    return handle


def main() -> int:
    # 子进程工作模式：由主进程 run_paddle_isolated 调起，处理一批后即退出。
    # 必须放在解析常规参数之前，因为 --_paddle_worker 不是用户可见的参数。
    if len(sys.argv) >= 3 and sys.argv[1] == "--_paddle_worker":
        return _paddle_worker_main(sys.argv[2])

    args = build_parser().parse_args()
    global RENDER_DPI, BINARIZE
    RENDER_DPI = args.dpi
    BINARIZE = args.binarize

    if args.check_env_only:
        print("[环境检查] OCR 隔离环境可用。")
        return 0

    if args.rebuild_results:
        raw_output_dir = args.output_dir
        if raw_output_dir is None:
            raw_output_dir = input(
                "请输入上一次 OCR 的结果保存文件夹路径（其中应有 .cma_ocr_resume）："
            )
        if not raw_output_dir or not raw_output_dir.strip():
            print("结果保存文件夹不能为空；未执行 OCR，也未修改任何结果。")
            return 2
        output_base = Path(clean_path_argument(raw_output_dir)).expanduser().resolve()
        if not output_base.is_dir():
            print(f"路径无效：{output_base}")
            return 2
        return rebuild_results_from_raw(output_base)

    raw_keywords = args.keywords
    if raw_keywords is None:
        raw_keywords = input(
            "请输入关键词（多个关键词建议用逗号分隔；含空格短语请加引号）："
        )
    keywords = parse_keywords(raw_keywords)

    raw_input_dir = args.input_dir
    if raw_input_dir is None:
        raw_input_dir = input("请输入包含 PDF 的根文件夹路径：")
    input_dir = Path(clean_path_argument(raw_input_dir)).expanduser().resolve()
    if not input_dir.is_dir():
        print(f"路径无效：{input_dir}")
        return 2

    raw_output_dir = args.output_dir
    if raw_output_dir is None:
        raw_output_dir = input(
            "请输入结果保存文件夹路径（直接回车则保存到当前文件夹）："
        )
    if raw_output_dir and raw_output_dir.strip():
        output_base = Path(clean_path_argument(raw_output_dir)).expanduser().resolve()
    else:
        output_base = Path.cwd().resolve()
    # 结果目录不存在则自动新建（含多级父目录）。
    output_base.mkdir(parents=True, exist_ok=True)
    try:
        run_lock = acquire_single_instance_lock(output_base)
    except RuntimeError as exc:
        print(f"[重复运行保护] {exc}")
        return 3
    # Keep the handle alive until process exit; the OS also releases it after a crash.
    _ = run_lock
    run_name = time.strftime("%Y%m%d_%H%M%S") + f"_{os.getpid()}"
    output_dir = output_base / "runs" / run_name
    configure_logging(output_dir, output_base)
    logging.info("关键词：%s", keywords)
    logging.info("输入目录：%s", input_dir)
    logging.info("输出目录：%s", output_dir)
    latest_pointer = output_base / "CMA_OCR_最近一次结果.txt"
    latest_pointer.write_text(str(output_dir), encoding="utf-8")

    pdf_files = discover_pdfs(input_dir)
    if not pdf_files:
        logging.error("没有找到 PDF 文件")
        return 1
    logging.info("发现 %d 个 PDF 文件", len(pdf_files))
    logging.info("渲染 DPI=%d | Tesseract二值化=%s", RENDER_DPI, BINARIZE)
    try:
        file_manifest = build_file_manifest(pdf_files)
    except OSError as exc:
        logging.exception("无法读取 PDF 文件清单信息：%s", exc)
        return 1
    errors: list[PageError] = []
    tesseract_stats: dict[str, int] | None = None
    paddle_stats: dict[str, int] | None = None

    # 续跑状态：崩溃后重跑时跳过已完成文件，只补没跑完的部分。
    state_dir = output_base / STATE_DIR_NAME
    state = _load_state(state_dir)
    same_job = (
        state.get("input_dir") == str(input_dir)
        and state.get("keywords") == keywords
        and state.get("dpi") == RENDER_DPI
    )
    # 旧版本状态没有文件清单。先安全升级、保留已有原始命中；从下一次起，
    # 任何 PDF 的新增、删除、替换或修改时间变化都会强制新任务，避免误跳过。
    if same_job and state.get("files") is None:
        state["files"] = file_manifest
        _save_state(state_dir, state)
        resume = True
        logging.warning(
            "检测到旧版续跑状态，已补写本次文件清单；本次保留原断点数据，"
            "后续文件变化将不再被错误跳过。"
        )
    else:
        resume = same_job and state.get("files") == file_manifest
    if not resume:
        state = {
            "input_dir": str(input_dir),
            "keywords": keywords,
            "dpi": RENDER_DPI,
            "files": file_manifest,
            "tesseract_done": [],
            "paddle_done": [],
        }
        _save_state(state_dir, state)
        (state_dir / "tesseract_raw.csv").unlink(missing_ok=True)
        (state_dir / "paddle_raw.csv").unlink(missing_ok=True)
        logging.info("未检测到可续跑状态，开始全新任务")
    else:
        logging.info(
            "检测到可续跑状态：tesseract 已完成 %d 个文件，paddle 已完成 %d 个文件",
            len(state.get("tesseract_done", [])),
            len(state.get("paddle_done", [])),
        )
    tesseract_raw_path = state_dir / "tesseract_raw.csv"
    paddle_raw_path = state_dir / "paddle_raw.csv"

    # 引擎分两个完整阶段顺序运行：各自打开 PDF、渲染页面、保存结果。
    # 两者不共享页面图像或 OCR 结果，合并前不分主次。
    tesseract_available = False
    try:
        configure_tesseract()
        tesseract_available = True
    except Exception as exc:
        errors.append(
            PageError(
                timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                engine="tesseract",
                filename="-",
                page="-",
                attempt="-",
                stage="引擎初始化",
                message=f"{type(exc).__name__}: {exc}",
            )
        )
        logging.exception("Tesseract 初始化失败，仅使用 PaddleOCR：%s", exc)

    if tesseract_available:
        _, tesseract_stats = run_single_engine(
            "tesseract", None, pdf_files, keywords, errors,
            tesseract_raw_path, state, state_dir,
        )
    write_error_csv(output_dir / f"{OUTPUT_PREFIX}_errors.csv", errors)
    gc.collect()

    # Paddle 阶段用独立子进程分批处理：每个子进程处理一批后退出，操作系统强制
    # 回收 Paddle 的 C++ 内存池（进程内 del/gc 无效），从而根治长期运行 OOM。
    paddle_stats, paddle_ran = run_paddle_isolated(
        pdf_files, keywords, errors, paddle_raw_path, state_dir,
        output_dir / f"{OUTPUT_PREFIX}_paddle_worker.log", PADDLE_BATCH_SIZE,
    )
    if not tesseract_available and not paddle_ran:
        logging.error("Tesseract 与 PaddleOCR 均不可用，跳过识别")

    write_error_csv(output_dir / f"{OUTPUT_PREFIX}_errors.csv", errors)

    # 从增量 raw CSV 读回全部命中（含续跑恢复的部分），再做最终合并。
    all_tesseract_hits = read_raw_hits(tesseract_raw_path)
    all_paddle_hits = read_raw_hits(paddle_raw_path)
    merged_hits = merge_and_deduplicate(all_tesseract_hits + all_paddle_hits)
    # B方案：主表只放可信命中；单引擎模糊命中单独进候选表，供人工复核。
    trusted_hits = [item for item in merged_hits if not item.is_candidate]
    candidate_hits = [item for item in merged_hits if item.is_candidate]

    merged_path = output_dir / f"{OUTPUT_PREFIX}_merged.csv"
    write_merged_csv(merged_path, trusted_hits)
    candidate_path = output_dir / f"{OUTPUT_PREFIX}_candidates.csv"
    write_candidate_csv(candidate_path, candidate_hits)
    error_path = output_dir / f"{OUTPUT_PREFIX}_errors.csv"
    write_error_csv(error_path, errors)
    write_raw_csv(output_dir / f"{OUTPUT_PREFIX}_tesseract_raw.csv", all_tesseract_hits)
    write_raw_csv(output_dir / f"{OUTPUT_PREFIX}_paddle_raw.csv", all_paddle_hits)

    # 每次的完整记录留在 runs\时间戳 子目录；同时把最新结果发布到
    # 主程序旁边，保持旧版启动脚本和用户习惯兼容。
    write_raw_csv(output_base / f"{OUTPUT_PREFIX}_tesseract_raw.csv", all_tesseract_hits)
    write_raw_csv(output_base / f"{OUTPUT_PREFIX}_paddle_raw.csv", all_paddle_hits)
    write_merged_csv(output_base / f"{OUTPUT_PREFIX}_merged.csv", trusted_hits)
    write_candidate_csv(output_base / f"{OUTPUT_PREFIX}_candidates.csv", candidate_hits)
    write_error_csv(output_base / f"{OUTPUT_PREFIX}_errors.csv", errors)

    try:
        write_results_xlsx(
            output_dir / f"{OUTPUT_PREFIX}.xlsx",
            trusted_hits,
            candidate_hits,
            all_tesseract_hits,
            all_paddle_hits,
            errors,
        )
        write_results_xlsx(
            output_base / f"{OUTPUT_PREFIX}.xlsx",
            trusted_hits,
            candidate_hits,
            all_tesseract_hits,
            all_paddle_hits,
            errors,
        )
    except Exception as exc:  # noqa: BLE001 - CSV 成果仍必须可交付
        errors.append(
            PageError(
                timestamp=time.strftime("%Y-%m-%d %H:%M:%S"),
                engine="output",
                filename="-",
                page="-",
                attempt="-",
                stage="导出 XLSX",
                message=f"{type(exc).__name__}: {exc}",
            )
        )
        logging.exception("XLSX 导出失败；CSV 结果仍已保留：%s", exc)
        write_error_csv(error_path, errors)
        write_error_csv(output_base / f"{OUTPUT_PREFIX}_errors.csv", errors)

    logging.info("Tesseract 原始命中：%d", len(all_tesseract_hits))
    logging.info("PaddleOCR 原始命中：%d", len(all_paddle_hits))
    logging.info("合并去重后命中：%d", len(merged_hits))
    logging.info("可信命中（主表）：%d", len(trusted_hits))
    logging.info("待复核候选（模糊）：%d", len(candidate_hits))
    logging.info("最终结果：%s", merged_path)
    logging.info("Excel 工作簿：%s", output_dir / f"{OUTPUT_PREFIX}.xlsx")
    logging.info("候选清单：%s", candidate_path)
    logging.info("错误明细：%s（%d 条）", error_path, len(errors))
    for name, stats in (("Tesseract", tesseract_stats), ("PaddleOCR", paddle_stats)):
        if stats is not None:
            logging.info(
                "[覆盖统计] %s | 已尝试=%d | 获得文本=%d | 空页=%d | "
                "页级失败=%d | 内存降级=%d",
                name, stats["pages"], stats["text"], stats["empty"],
                stats["error"], stats["fallback"],
            )
    logging.info("本次完整输出目录：%s", output_dir)

    # 全部文件都处理完后清除续跑状态；否则保留，供下次断点续跑。
    # 重新从磁盘读回状态（Paddle 阶段由子进程更新 paddle_done，主进程 dict 未同步）。
    state = _load_state(state_dir)
    if (
        len(state.get("tesseract_done", [])) >= len(pdf_files)
        and len(state.get("paddle_done", [])) >= len(pdf_files)
    ):
        _clear_state(state_dir)
        logging.info("全部文件处理完成，续跑状态已清除")
    else:
        logging.warning(
            "本次未处理完全部文件（tesseract=%d/%d, paddle=%d/%d）；"
            "续跑状态已保留，下次运行将从此继续",
            len(state.get("tesseract_done", [])), len(pdf_files),
            len(state.get("paddle_done", [])), len(pdf_files),
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())

